import os
import re
import pandas as pd
import openpyxl

class PreProcessing:
    def __init__(
        self,
        file_in: str = "data_pre.xlsx",
        file_clean: str = "data_clean.xlsx",
        file_blocks: str = "bloques.xlsx",
        file_template: str = "LFR Plantilla.xlsx",
        file_out: str = "data.xlsx",
        save_intermediate: bool = False,
        cleanup_intermediate: bool = True,
    ) -> None:
        self.file_in = file_in
        self.file_clean = file_clean
        self.file_blocks = file_blocks
        self.file_template = file_template
        self.file_out = file_out
        self.save_intermediate = save_intermediate
        self.cleanup_intermediate = cleanup_intermediate

        # buffers in-memory cuando save_intermediate=False
        self.cleaned_sheets = None           # dict[str, DataFrame]
        self.blocks_full_df = {}             # dict[str, DataFrame] (incluye 2 filas de header + cuerpo)
        self.blocks_headers = {}             # dict[str, list[list]]  (dos filas)
    
    # ------------------------------
    # Helpers
    # ------------------------------
    @staticmethod
    def _is_blank_cell(x):
        return pd.isna(x) or (isinstance(x, str) and x.strip() == "")

    @staticmethod
    def _is_empty_row(row) -> bool:
        # Fila separadora = TODAS sus celdas vacías
        return all(PreProcessing._is_blank_cell(v) for v in row)

    @staticmethod
    def split_by_empty_rows(df: pd.DataFrame) -> list:
        """Parte df en subsecciones cada vez que encuentra una fila totalmente vacía."""
        if df is None or df.empty:
            return []
        sections, cur = [], []
        for _, row in df.iterrows():
            if PreProcessing._is_empty_row(row):
                if cur:
                    sections.append(pd.DataFrame(cur, columns=df.columns))
                    cur = []
            else:
                cur.append(row.tolist())
        if cur:
            sections.append(pd.DataFrame(cur, columns=df.columns))
        return sections

    @staticmethod
    def ensure_top_stub_row(df: pd.DataFrame, suffix: str = "1") -> pd.DataFrame:
        """
        Inserta una fila stub al inicio SOLO si la primera fila NO es stub.
        Definición de stub: A con valor y B vacía (ignoramos el resto de columnas).
        La nueva stub será: A = <donor> + suffix (p. ej. 'Caja1'), B.. = NA.
        El donor por defecto es A de la primera fila (si está vacía, busca la primera A no vacía; si no hay, 'Sección').
        """
        if df is None or df.empty:
            return df

        ncols = df.shape[1]
        first_a_has = not PreProcessing._is_blank_cell(df.iat[0, 0])
        first_b_blank = True if ncols < 2 else PreProcessing._is_blank_cell(df.iat[0, 1])
        if first_a_has and first_b_blank:
            return df

        # Donor
        if not PreProcessing._is_blank_cell(df.iat[0, 0]):
            donor = str(df.iat[0, 0]).strip()
        else:
            donor = None
            for r in range(df.shape[0]):
                if not PreProcessing._is_blank_cell(df.iat[r, 0]):
                    donor = str(df.iat[r, 0]).strip()
                    break
            if donor is None:
                donor = "Sección"

        stub_name = f"{donor}{suffix}"
        stub = pd.DataFrame([[stub_name] + [pd.NA] * (ncols - 1)], columns=df.columns)
        return pd.concat([stub, df], ignore_index=True)

    @staticmethod
    def apply_stub_to_all_subsections(df: pd.DataFrame, suffix: str = "1") -> pd.DataFrame:
        """
        Aplica ensure_top_stub_row a cada subsección (delimitada por filas completamente vacías)
        y recompone dejando UNA sola fila vacía entre subsecciones.
        """
        if df is None or df.empty:
            return df
        sections = PreProcessing.split_by_empty_rows(df)
        if not sections:
            return df

        processed = []
        for sec in sections:
            if sec is None or sec.empty:
                continue
            sec2 = PreProcessing.ensure_top_stub_row(sec, suffix=suffix)
            processed.append(sec2)

        if not processed:
            return pd.DataFrame(columns=df.columns)

        out_parts = []
        for k, sec in enumerate(processed):
            out_parts.append(sec)
            if k < len(processed) - 1:
                out_parts.append(pd.DataFrame([[None] * df.shape[1]], columns=df.columns))  # separador

        return pd.concat(out_parts, ignore_index=True)

    @staticmethod
    def _split_sections_measures(df: pd.DataFrame) -> list:
        """
        Lógica para partir secciones desde 'measures' y cerrar al hallar 2 filas vacías.
        Devuelve lista de DataFrames recortados para quitar las 2 primeras filas ([2:]).
        """
        sections = []
        current_section = []
        empty_count = 0
        in_section = False

        for idx, val in df.iloc[:, 0].items():
            if isinstance(val, str) and "measures" in val.lower():
                if current_section:
                    if len(current_section) > 2:
                        sections.append(pd.DataFrame(current_section[2:]).reset_index(drop=True))
                    current_section = []
                in_section = True
                empty_count = 0

            if in_section:
                current_section.append(df.iloc[idx].tolist())

                if pd.isna(val) or (isinstance(val, str) and val.strip() == ""):
                    empty_count += 1
                    if empty_count >= 2:
                        if len(current_section) > 2:
                            sections.append(pd.DataFrame(current_section[:-2][2:]).reset_index(drop=True))
                        current_section = []
                        in_section = False
                        empty_count = 0
                else:
                    empty_count = 0

        if current_section and len(current_section) > 2:
            sections.append(pd.DataFrame(current_section[2:]).reset_index(drop=True))

        return sections

    @staticmethod
    def _insert_empty_before_total_aconten(bloque: pd.DataFrame) -> pd.DataFrame:
        for j in range(len(bloque)):
            if isinstance(bloque.iloc[j, 0], str) and bloque.iloc[j, 0].strip().upper() == "TOTAL ACONTEN":
                empty_row = pd.DataFrame([[None] * bloque.shape[1]], columns=bloque.columns)
                bloque = pd.concat([bloque.iloc[:j], empty_row, bloque.iloc[j:]], ignore_index=True)
                break
        return bloque

    # ------------------------------
    # PARTE 1
    # ------------------------------
    def build_data_clean(self):
        sheets = pd.read_excel(self.file_in, sheet_name=None, header=None)
        cleaned_sheets = {}

        for sheet_name, df in sheets.items():
            df = df.copy()
            rows_to_drop = []

            for idx, val in df.iloc[:, 0].items():  # recorrer primera columna
                if isinstance(val, str) and "measures" in val.lower():
                    # ir hacia arriba hasta encontrar fila vacía
                    up = idx - 1
                    while up >= 0:
                        v = df.iloc[up, 0]
                        if pd.isna(v) or (isinstance(v, str) and v.strip() == ""):
                            break
                        rows_to_drop.append(up)
                        up -= 1

            rows_to_drop = sorted(set(rows_to_drop))

            # DEBUG
            if rows_to_drop:
                print(f"\n[{sheet_name}] Se eliminarán {len(rows_to_drop)} filas:")
                for r in rows_to_drop:
                    valor = df.iloc[r, 0]
                    print(f"  - Fila Excel {r+1} (índice {r}) → {repr(valor)}")
            else:
                print(f"\n[{sheet_name}] No se eliminarán filas.")

            df = df.drop(rows_to_drop).reset_index(drop=True)
            cleaned_sheets[sheet_name] = df

        # Guardar o mantener en memoria
        if self.save_intermediate:
            with pd.ExcelWriter(self.file_clean, engine="openpyxl") as writer:
                for sheet_name, df in cleaned_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            print(f"\nArchivo guardado en {self.file_clean}")
        else:
            self.cleaned_sheets = cleaned_sheets  # mantener en memoria

        # devolver header de 'General' (primeras 2 filas) para usarlo en bloques
        if "General" in cleaned_sheets and not cleaned_sheets["General"].empty:
            gen = cleaned_sheets["General"]
            if len(gen) >= 2:
                header2 = gen.iloc[:2].fillna("").values.tolist()
            elif len(gen) == 1:
                header2 = [gen.iloc[0].fillna("").tolist(), [""] * gen.shape[1]]
            else:
                header2 = [[""], [""]]
        else:
            header2 = [[""], [""]]
        return cleaned_sheets, header2

    # ------------------------------
    # PARTE 2
    # ------------------------------
    def build_bloques(self, df_clean_header2):
        file_in = self.file_in

        # Procesar hoja "General" de data_pre.xlsx (tal cual original)
        df = pd.read_excel(file_in, sheet_name="General", header=None)
        sections = self._split_sections_measures(df)

        bloques = {
            "Total Lima + 13 Ciudades": [],
            "Total Lima": [],
            "Total 13 Ciudades": []
        }

        for sec in sections[:-1]:
            col0 = sec.iloc[:, 0]
            i = 0
            while i < len(sec):
                val = col0.iloc[i]
                if isinstance(val, str) and val.strip() in bloques:
                    start = i
                    i += 1
                    while i < len(sec) and not pd.isna(col0.iloc[i]) and str(col0.iloc[i]).strip() != "":
                        i += 1
                    end = i
                    bloque = sec.iloc[start:end].reset_index(drop=True)
                    bloque = self._insert_empty_before_total_aconten(bloque)
                    empty_row = pd.DataFrame([[None] * bloque.shape[1]], columns=bloque.columns)
                    bloque_con_espacio = pd.concat([bloque, empty_row], ignore_index=True)
                    bloques[val.strip()].append(bloque_con_espacio)
                i += 1

        # última sección completa
        if sections:
            ultima_sec = sections[-1]
            empty_row = pd.DataFrame([[None] * ultima_sec.shape[1]], columns=ultima_sec.columns)
            ultima_sec_con_espacio = pd.concat([ultima_sec.reset_index(drop=True), empty_row], ignore_index=True)
            bloques["Total Lima + 13 Ciudades"].append(ultima_sec_con_espacio)

        for k in bloques:
            if bloques[k]:
                bloques[k] = pd.concat(bloques[k], ignore_index=True)
            else:
                bloques[k] = pd.DataFrame()

        # escribir en workbook (disco o memoria), siempre aplicando stub
        # Renombrar hojas luego
        rename_map = {
            "Total Lima + 13 Ciudades": "General",
            "Total Lima": "Lima",
            "Total 13 Ciudades": "Provincias 13"
        }

        if self.save_intermediate:
            # Paso 1: crear archivo con hojas vacías
            with pd.ExcelWriter(self.file_blocks, engine="openpyxl") as writer:
                for name in bloques.keys():
                    pd.DataFrame().to_excel(writer, sheet_name=name, index=False, header=False)
            wb = openpyxl.load_workbook(self.file_blocks)

            # Escribir contenido + header
            for name, df_out in bloques.items():
                ws = wb[name]
                # headers (A1, A2)
                for r_idx, row in enumerate(df_clean_header2, start=1):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                # stub subsecciones
                df_out = self.apply_stub_to_all_subsections(df_out, suffix="1")
                # A3 en adelante
                for r_idx, row in enumerate(df_out.values, start=3):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

            # Renombrar
            for old_name, new_name in rename_map.items():
                if old_name in wb.sheetnames:
                    wb[old_name].title = new_name

            # Otras hojas
            xls = pd.ExcelFile(file_in)
            # Usar el mismo header de df_clean_header2 para todas (como en tu flujo)
            for sheet in xls.sheet_names:
                if sheet == "General":
                    continue
                df_sheet = pd.read_excel(file_in, sheet_name=sheet, header=None)
                sections_other = self._split_sections_measures(df_sheet)
                if sections_other:
                    secciones_con_espacios = []
                    for sec in sections_other:
                        secciones_con_espacios.append(sec)
                        empty_row = pd.DataFrame([[None] * sec.shape[1]], columns=sec.columns)
                        secciones_con_espacios.append(empty_row)
                    df_concat = pd.concat(secciones_con_espacios, ignore_index=True)
                else:
                    df_concat = pd.DataFrame()

                if sheet not in wb.sheetnames:
                    ws = wb.create_sheet(title=sheet)
                else:
                    ws = wb[sheet]

                # headers
                for r_idx, row in enumerate(df_clean_header2, start=1):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

                # stub subsecciones
                df_concat = self.apply_stub_to_all_subsections(df_concat, suffix="1")
                # A3
                for r_idx, row in enumerate(df_concat.values, start=3):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

            wb.save(self.file_blocks)
            print(f"Archivo final con todas las hojas creado: {self.file_blocks}")

        else:
            # Construir "bloques" completamente en memoria como DataFrame por hoja (incluye header 2 filas)
            # Primero las tres especiales con rename_map
            for old_name, df_body in bloques.items():
                new_name = rename_map.get(old_name, old_name)
                df_body = self.apply_stub_to_all_subsections(df_body, suffix="1")
                # Determinar ancho
                max_cols = max(
                    df_body.shape[1] if not df_body.empty else 1,
                    max(len(r) for r in df_clean_header2) if df_clean_header2 else 1
                )
                # Extender headers a max_cols
                headers = []
                for r in df_clean_header2:
                    row_ext = list(r) + [""] * (max_cols - len(r))
                    headers.append(row_ext)
                # Asegurar ancho del cuerpo
                if not df_body.empty and df_body.shape[1] < max_cols:
                    for _ in range(max_cols - df_body.shape[1]):
                        df_body[df_body.shape[1]] = None
                # DataFrame final
                df_full = pd.DataFrame(headers + df_body.fillna("").values.tolist())
                self.blocks_full_df[new_name] = df_full
                self.blocks_headers[new_name] = headers

            # Otras hojas del archivo de entrada
            xls = pd.ExcelFile(file_in)
            for sheet in xls.sheet_names:
                if sheet == "General":
                    continue
                df_sheet = pd.read_excel(file_in, sheet_name=sheet, header=None)
                sections_other = self._split_sections_measures(df_sheet)
                if sections_other:
                    secciones_con_espacios = []
                    for sec in sections_other:
                        secciones_con_espacios.append(sec)
                        empty_row = pd.DataFrame([[None] * sec.shape[1]], columns=sec.columns)
                        secciones_con_espacios.append(empty_row)
                    df_concat = pd.concat(secciones_con_espacios, ignore_index=True)
                else:
                    df_concat = pd.DataFrame()
                df_concat = self.apply_stub_to_all_subsections(df_concat, suffix="1")

                # Determinar ancho
                max_cols = max(
                    df_concat.shape[1] if not df_concat.empty else 1,
                    max(len(r) for r in df_clean_header2) if df_clean_header2 else 1
                )
                # Extender headers a max_cols
                headers = []
                for r in df_clean_header2:
                    row_ext = list(r) + [""] * (max_cols - len(r))
                    headers.append(row_ext)
                # Asegurar ancho del cuerpo
                if not df_concat.empty and df_concat.shape[1] < max_cols:
                    for _ in range(max_cols - df_concat.shape[1]):
                        df_concat[df_concat.shape[1]] = None

                df_full = pd.DataFrame(headers + (df_concat.fillna("").values.tolist() if not df_concat.empty else []))
                self.blocks_full_df[sheet] = df_full
                self.blocks_headers[sheet] = headers

        return True

    # ------------------------------
    # PARTE 3
    # ------------------------------
    @staticmethod
    def normalize_key(x):
        """Normalizar texto: None -> '', strip, colapsar espacios, lower."""
        if pd.isna(x):
            return ""
        s = str(x).strip()
        s = re.sub(r"\s+", " ", s)
        return s.lower()

    @staticmethod
    def extract_sections_from_df(df, start_row=2):
        """
        Extrae secciones a partir de la fila start_row (0-indexed).
        Devuelve lista de DataFrames (cada sección mantiene todas las columnas).
        Se separa cada vez que la columna A es vacía.
        """
        rows = df.iloc[start_row:].reset_index(drop=True)
        sections = []
        cur = []
        for _, row in rows.iterrows():
            val = row.iloc[0]
            if pd.isna(val) or str(val).strip() == "":
                if cur:
                    sections.append(pd.DataFrame(cur, columns=df.columns))
                    cur = []
            else:
                cur.append(row.tolist())
        if cur:
            sections.append(pd.DataFrame(cur, columns=df.columns))
        return sections

    @staticmethod
    def align_section_by_template(template_keys, block_df):
        """
        Reordena block_df conservando solo filas cuyo first-col normalizada
        aparece en template_keys (lista de normalized keys) y en el orden de template_keys.
        """
        if block_df.empty:
            return block_df.copy()

        block_keys = [PreProcessing.normalize_key(x) for x in block_df.iloc[:, 0].tolist()]
        mapping = {}
        for i, k in enumerate(block_keys):
            mapping.setdefault(k, []).append(i)

        new_rows = []
        for t in template_keys:
            lst = mapping.get(t)
            if lst:
                idx = lst.pop(0)  # primera aparición
                new_rows.append(block_df.iloc[idx].tolist())

        if not new_rows:
            return pd.DataFrame(columns=block_df.columns)
        return pd.DataFrame(new_rows, columns=block_df.columns)

    def build_data_from_template(self):
        # Cargar plantilla
        xls_template = pd.ExcelFile(self.file_template)
        template_sections_norm = {}  # sheet -> list[list[str]]
        template_sections_raw  = {}  # sheet -> list[DataFrame]

        for sheet in xls_template.sheet_names:
            df_t = pd.read_excel(self.file_template, sheet_name=sheet, header=None)
            secs = self.extract_sections_from_df(df_t, start_row=2)
            template_sections_raw[sheet] = secs
            template_sections_norm[sheet] = [
                [self.normalize_key(v) for v in sec.iloc[:,0].tolist()] for sec in secs
            ]

        # Cargar bloques: desde disco o memoria
        blocks_headers = {}
        blocks_sections = {}
        sheet_names = []

        if self.save_intermediate:
            xls_blocks = pd.ExcelFile(self.file_blocks)
            sheet_names = xls_blocks.sheet_names
            for sheet in sheet_names:
                df_b = pd.read_excel(self.file_blocks, sheet_name=sheet, header=None)
                # headers
                if len(df_b) >= 2:
                    blocks_headers[sheet] = df_b.iloc[:2].fillna("").values.tolist()
                elif len(df_b) == 1:
                    blocks_headers[sheet] = [df_b.iloc[0].fillna("").tolist(), [""] * df_b.shape[1]]
                else:
                    blocks_headers[sheet] = [[""], [""]]
                # secciones
                blocks_sections[sheet] = self.extract_sections_from_df(df_b, start_row=2)
        else:
            sheet_names = list(self.blocks_full_df.keys())
            for sheet in sheet_names:
                df_b = self.blocks_full_df[sheet]
                blocks_headers[sheet] = self.blocks_headers.get(sheet, [[""], [""]])
                blocks_sections[sheet] = self.extract_sections_from_df(df_b, start_row=2)

        # Alinear y construir secciones finales
        aligned_sections_by_sheet = {}

        for sheet in sheet_names:
            sections = blocks_sections.get(sheet, [])
            final_sections = []
            template_secs_for_sheet_norm = template_sections_norm.get(sheet, None)
            template_secs_for_sheet_raw  = template_sections_raw.get(sheet, None)

            if template_secs_for_sheet_norm:
                for i, blk_sec in enumerate(sections):
                    if i < len(template_secs_for_sheet_norm):
                        t_keys = template_secs_for_sheet_norm[i]

                        # Reemplazar segunda fila (col A) con título real de plantilla (si existe)
                        if not blk_sec.empty and blk_sec.shape[0] >= 2 and template_secs_for_sheet_raw and i < len(template_secs_for_sheet_raw):
                            plantilla_real_title = template_secs_for_sheet_raw[i].iloc[0, 0]
                            blk_sec = blk_sec.copy()
                            blk_sec.iat[1, 0] = plantilla_real_title

                        aligned = self.align_section_by_template(t_keys, blk_sec)
                        final_sections.append(aligned)
                    else:
                        final_sections.append(blk_sec)
            else:
                final_sections = sections.copy()

            # Filler al inicio de cada sección (misma lógica efectiva que usabas)
            new_sections = []
            for sec in final_sections:
                if sec.empty:
                    new_sections.append(sec)
                    continue
                val = sec.iat[0, 0]
                filler = pd.DataFrame([[val] + [pd.NA] * (sec.shape[1] - 1)], columns=sec.columns)
                sec = pd.concat([filler, sec], ignore_index=True)
                new_sections.append(sec)

            aligned_sections_by_sheet[sheet] = new_sections

        # Escribir workbook final data.xlsx
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        for sheet in sheet_names:
            ws = wb.create_sheet(title=sheet)

            # headers
            headers = blocks_headers.get(sheet, [[""], [""]])
            max_cols = 0
            for r in headers:
                max_cols = max(max_cols, len(r))
            for sec in aligned_sections_by_sheet.get(sheet, []):
                if not sec.empty:
                    max_cols = max(max_cols, sec.shape[1])
            if max_cols == 0:
                max_cols = 1

            # A1/A2
            for r_idx, row in enumerate(headers, start=1):
                row_ext = list(row) + [""] * (max_cols - len(row))
                for c_idx, val in enumerate(row_ext, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=("" if pd.isna(val) else val))

            # A3 en adelante
            write_row = 3
            for sec in aligned_sections_by_sheet.get(sheet, []):
                if sec.empty:
                    write_row += 1
                    continue
                if sec.shape[1] < max_cols:
                    for _ in range(max_cols - sec.shape[1]):
                        sec[sec.shape[1]] = None
                for _, row in sec.iterrows():
                    for c_idx in range(max_cols):
                        val = row.iloc[c_idx] if c_idx < len(row) else None
                        ws.cell(row=write_row, column=c_idx+1, value=(None if pd.isna(val) else val))
                    write_row += 1
                write_row += 1  # 1 fila vacía entre secciones

        wb.save(self.file_out)
        print("Proceso finalizado. Archivo generado:", self.file_out)

        # Limpieza de intermedios si se pidieron y se guardaron
        if self.save_intermediate and self.cleanup_intermediate:
            for path in (self.file_clean, self.file_blocks):
                try:
                    if os.path.exists(path):
                        os.remove(path)
                        print(f"Intermedio eliminado: {path}")
                except Exception as e:
                    print(f"Advertencia: no se pudo eliminar {path}: {e}")



    def run(self) -> None:
        _, header2 = self.build_data_clean()
        self.build_bloques(header2)
        self.build_data_from_template()
