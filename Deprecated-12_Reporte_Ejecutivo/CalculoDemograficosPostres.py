import Funciones_Formato
import pandas as pd
import re
from openpyxl import load_workbook
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import random
import pandas as pd
import re
import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import random


class DatosDemograficoPostres:
    def __init__(self, df, filename, nombre_hoja):
        self.df = df
        self.filename = filename
        self.nombre_hoja = nombre_hoja
        self.messures = {
            "Weighted R_VOL1 Vert %": "% Volumen",
            "Weighted PENET": "Penetración (%)",
            "Weighted VO1_BUY": "Compra media (kg)",
            "Weighted VO1_DAY": "Compra por acto (kg)", 
            "Weighted FREQ": "Frecuencia (veces)"
        }
        self.indices = {
            "WPOBRPO = T. Pos + Fla + Gel + Azl\\Total WPOBRPO": "Industria",		
            "WPOBRPO = T. Danone\\T. Pos + Fla + Gel + Azl\\Total WPOBRPO": "Danone",			
            "WPOBRPO = T. Danette\\T. Danone\\T. Pos + Fla + Gel + Azl\\Total WPOBRPO": "Yogurisimo",
            "WPOBRPO = T. Ser\\T. Danone\\T. Pos + Fla + Gel + Azl\\Total WPOBRPO": "SER",
            "WPOBRPO = T. Serenito\\T. Danone\\T. Pos + Fla + Gel + Azl\\Total WPOBRPO": "Serenito"
        }

    def _obtener_dataframes_por_indices(self, dataframes, palabras_a_buscar):
        cadenas_presentes = [cadena for cadena in palabras_a_buscar if cadena in dataframes.columns]
        posiciones = {cadena: dataframes.columns.get_loc(cadena) for cadena in cadenas_presentes}
        return dict(sorted(posiciones.items(), key=lambda item: item[1]))

    def _ordenar_df(self):
        posiciones = self._obtener_dataframes_por_indices(self.df, self.indices)
        dataframes = {}
        medidas = list(posiciones.keys())
        for indice, medida in enumerate(medidas):
            posicion_inicio = posiciones[medida]
            if indice < len(medidas) - 1:
                siguiente_region = medidas[indice + 1]
                siguiente_posicion_inicio = posiciones[siguiente_region]
                posicion_fin = siguiente_posicion_inicio - 1
            else:
                posicion_fin = self.df.shape[1]

            titulo = self.indices[medida]
            dataframes[titulo] = self.df.iloc[:, posicion_inicio:posicion_fin + 1]
            dataframes[titulo] = dataframes[titulo].dropna(axis=1, how='all')

        return dataframes

    def _obtener_dataframes_por_titulo_y_region(self):
        def _obtener_dataframes_por_titulo_y_region_interno(data, titulo_a_buscar, palabras_a_buscar):
            primera_columna = data[titulo_a_buscar].iloc[:, 0]
            posiciones_regiones = {}
            for palabra in palabras_a_buscar:
                pattern = re.compile(re.escape(palabra), re.IGNORECASE)
                matches = primera_columna.apply(lambda x: bool(pattern.search(str(x))))
                if any(matches):
                    posiciones_regiones[palabra] = matches.idxmax()
                else:
                    posiciones_regiones[palabra] = None

            dataframes_regiones_separados = {}
            regiones = list(posiciones_regiones.keys())
            for indice, region in enumerate(regiones):
                posicion_inicio = posiciones_regiones[region]

                if indice < len(regiones) - 1:
                    siguiente_region = regiones[indice + 1]
                    siguiente_posicion_inicio = posiciones_regiones[siguiente_region]
                    posicion_fin = siguiente_posicion_inicio - 1
                else:
                    posicion_fin = data[titulo_a_buscar].shape[0] - 1

                nombre = self.messures[region]
                dataframes_regiones_separados[nombre] = data[titulo_a_buscar].iloc[posicion_inicio:posicion_fin + 1]
                dataframes_regiones_separados[nombre] = dataframes_regiones_separados[nombre].dropna(axis=0, how='all')
                dataframes_regiones_separados[nombre] = dataframes_regiones_separados[nombre].dropna()

            return dataframes_regiones_separados

        data = self._ordenar_df()
        dataframes_finales = {}

        for titulo in self.indices.values():
            palabras_a_buscar = list(self.messures.keys())
            ingreso_datos = _obtener_dataframes_por_titulo_y_region_interno(data, titulo, palabras_a_buscar)
            dataframes_finales[titulo] = ingreso_datos

        return dataframes_finales

    def _cacular_promedio_geograficamente(self, dataframe):
        nuevos_dataframes = {}
        recibir_ytd = 12
        recibir_tres_meses = 3
        recibir_cuadtro_meses = 4

        for indicador in self.indices.values():
            new_dataframe = {}

            for medida in self.messures.values():
                primera_columna = dataframe[indicador][medida].iloc[:, 0]

                ultimos_4_actual = dataframe[indicador][medida].iloc[:, -(recibir_cuadtro_meses):]
                promedio_4_actual = ultimos_4_actual.mean(axis=1) + 0.0000000000001
                ult_4_anteriores = dataframe[indicador][medida].iloc[:, -(recibir_cuadtro_meses + recibir_cuadtro_meses):-(recibir_cuadtro_meses)]
                promedio_4_anteriores = ult_4_anteriores.mean(axis=1) + 0.0000000000001

                ultimos_mes_actual = dataframe[indicador][medida].iloc[:, -(1):]
                promedio_mes_actual = ultimos_mes_actual.mean(axis=1) + 0.0000001
                ult_mes_anteriores = dataframe[indicador][medida].iloc[:, -(recibir_cuadtro_meses + 1):-(recibir_cuadtro_meses)]
                promedio_mes_anteriores = ult_mes_anteriores.mean(axis=1) + 0.0000001

                if medida == "% Volumen":
                    nuevo_dataframe = pd.DataFrame({
                        ('', f"{medida}"): primera_columna,
                        (f"{indicador}   ", 'P3M (últ trimestre movil)'): promedio_mes_actual,
                        (f"{indicador}   ", 'Dif vs PY  '): (promedio_mes_actual - promedio_mes_anteriores),
                        (f"{indicador}  ", '  '): None,
                        (f"{indicador}   ", 'P12M (Promedio últ 4 P3M)'): promedio_4_actual,
                        (f"{indicador}    ", 'Dif vs PY '): (promedio_4_actual - promedio_4_anteriores), 
                    })
                elif medida == "Penetración (%)":
                    nuevo_dataframe = pd.DataFrame({
                        f"{medida}": primera_columna,
                        'P3M (últ trimestre movil)': promedio_mes_actual,
                        'Dif vs PY  ': (promedio_mes_actual - promedio_mes_anteriores),
                        '': None,
                        'P12M (Promedio últ 4 P3M)': promedio_4_actual,
                        'Dif vs PY ': (promedio_4_actual - promedio_4_anteriores),
                        })
                else:
                    try:
                        variacion_mes_actual = promedio_mes_actual / promedio_mes_anteriores - 1
                    except ZeroDivisionError:
                        variacion_mes_actual = 0

                    try:
                        variacion_4 = promedio_4_actual / promedio_4_anteriores - 1
                    except ZeroDivisionError:
                        variacion_4 = 0

                    nuevo_dataframe = pd.DataFrame({
                        f"{medida}": primera_columna,
                        'P3M (últ trimestre movil)': promedio_mes_actual,
                        '   %Variacion vs PY': variacion_mes_actual,
                        '': None,
                        'P12M (Promedio últ 4 P3M)': promedio_4_actual,
                        '%Variacion vs PY ': variacion_4,
                        })
                    
                new_dataframe[medida] = nuevo_dataframe
            nuevos_dataframes[indicador] = new_dataframe

        return nuevos_dataframes

    def _escribir_en_excel(self, dataframes):
        wb = load_workbook(filename=self.filename)
        ws = wb.create_sheet(self.nombre_hoja)

        max_col = 0 
        for indicador in self.indices.values():
            max_col = max_col + 0
            last_row = 5
            for medida in self.messures.values():
                rows = dataframe_to_rows(dataframes[indicador][medida], index=False, header=True)
                for r_idx, row in enumerate(rows, 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=last_row + r_idx, column=c_idx + max_col, value=value)

                last_row = last_row + r_idx + 1

            ws.cell(row=1, column=ws.max_column + 1, value="")
            max_col = ws.max_column

        Funciones_Formato.aplicar_formato_geo(ws,self.nombre_hoja)
        
        for fila in ws.iter_rows(min_row=1, max_row=10):
            # Acceder a las celdas de la fila
            if any(celda.value for celda in fila):
                Funciones_Formato.merge_cells_in_column(ws, fila[0].row)
        celdas_combinadas = Funciones_Formato.encontrar_celdas_combinadas_en_filas(ws, range(1, 11))
        Funciones_Formato.aplicar_relleno_y_color_a_celdas_combinadas(ws, celdas_combinadas)
        Funciones_Formato.Aplicar_nombres_relleno(ws,self.nombre_hoja)
        wb.save(self.filename)


    def procesar(self):
        diccionario_datos = self._obtener_dataframes_por_titulo_y_region()
        BD = self._cacular_promedio_geograficamente(diccionario_datos)
        self._escribir_en_excel(BD)

#nombre_excel = "a3.xlsx" ## nombre del archivo final
#directorio_actual = os.path.dirname(os.path.abspath(__file__))
#direccion_base1 = os.path.join(directorio_actual, "Datos_05.xlsx")
#df = pd.read_excel(direccion_base1,sheet_name="Demo_PO")
#nombre_excel = "Libro1.xlsx"
#processor =CalculoDemograficoPostres(df,nombre_excel, "Demo_PO")
#processor.procesar()
