import Funciones_Formato
import pandas as pd
import re
from openpyxl import load_workbook
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
import os

class Realizar_hoja_formato_manual:  #SegmentoCanalesProcessor
    def __init__(self, df, filename, nombre_hoja):
        self.df = df
        self.filename = filename
        self.nombre_hoja = nombre_hoja
        self.messures = {
            "Weighted PENET": "Penetración (%)", 
            "Weighted VO1_BUY": "Compra media (kg)",
            "Weighted VO1_DAY": "Compra por acto (kg)", 
            "Weighted FREQ": "Frecuencia (veces)"
        }
    
    def _obtener_dataframes_por_titulo_y_region(self):
        primera_columna = self.df.iloc[:, 0]
        posiciones_regiones = {}
        for palabra in self.messures.keys():
            pattern = re.compile(re.escape(palabra), re.IGNORECASE)
            matches = primera_columna.apply(lambda x: bool(pattern.search(str(x))))
            posiciones_regiones[palabra] = matches.idxmax() if any(matches) else None
        return dict(sorted(posiciones_regiones.items(), key=lambda item: item[1]))

    def _encontrar_fila_con_fechas(self):
        pattern = r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2}"
        numero_de_fila_con_fechas = None
        for indice, fila in self.df.iterrows():
            if any(fila.astype(str).str.match(pattern)):
                numero_de_fila_con_fechas = indice
                break
        dfdatos = self.df.iloc[numero_de_fila_con_fechas, :].tolist()
        dfdatos = [x.strftime('%b-%y') if isinstance(x, datetime.datetime) else x for x in dfdatos]
        return dfdatos

    def _ordenar_df(self):
        posiciones = self._obtener_dataframes_por_titulo_y_region()
        dataframes = {}
        medidas = list(posiciones.keys())
        for indice, medida in enumerate(medidas):
            posicion_inicio = posiciones[medida]
            if indice < len(medidas) - 1:
                siguiente_region = medidas[indice + 1]
                siguiente_posicion_inicio = posiciones[siguiente_region]
                posicion_fin = siguiente_posicion_inicio - 1
            else:
                posicion_fin = self.df.shape[0] - 1
            titulo = self.messures[medida]
            dataframes[titulo] = self.df.iloc[posicion_inicio:posicion_fin + 1]
            dataframes[titulo] = dataframes[titulo].dropna(axis=0, how='all')
        
        fechas = self._encontrar_fila_con_fechas()
        nuevos_dataframes = {}
        for clave, df in dataframes.items():
            df_sin_nan = df.dropna()
            df_sin_nan.columns = fechas
            filas_con_ceros = (df_sin_nan == 0).sum(axis=1)
            filas_a_eliminar = filas_con_ceros[filas_con_ceros > 11].index
            df_filtrado = df_sin_nan.drop(filas_a_eliminar)
            nuevos_dataframes[clave] = df_filtrado 
        return nuevos_dataframes

    def _separar_fechas_YTD(self, variable): 
        elementos = str(variable).split('-')
        YTD_mes, YTD_año = elementos[0], elementos[1]
        numero_año_anterior = int(YTD_año) - 1
        return YTD_mes, YTD_año, str(numero_año_anterior)

    def _calcular_promedio_ultimos_12_meses(self, dataframes):
        nuevos_dataframes = {}
        recibir_ytd = 12
        recibir_tres_meses = 3
        for medida in self.messures.values(): 
            primera_columna = dataframes[medida].iloc[:, 0]
            ultimos_meses_año_actual = dataframes[medida].iloc[:, -recibir_ytd:]
            promedio_año_actual = ultimos_meses_año_actual.mean(axis=1) + 0.0000000000001
            ult_meses_promedio_PY = dataframes[medida].iloc[:, -(recibir_ytd + recibir_ytd):-recibir_ytd]
            promedio_meses_promedio_PY = ult_meses_promedio_PY.mean(axis=1) + 0.0000000000001

            ultima_columna = dataframes[medida].columns[-1]
            YTD_mes, YTD_año, numero_año_antes = self._separar_fechas_YTD(ultima_columna)

            ultimos_YTD_actual = dataframes[medida].loc[:, ('Jan-' + YTD_año):(ultima_columna)]
            promedio_YTD_ctual = ultimos_YTD_actual.mean(axis=1)+0.000000001
            ult_YTD_anterior = dataframes[medida].loc[:, ('Jan-' + numero_año_antes):(YTD_mes + '-' + numero_año_antes)]
            promedio_YTD_anterior = ult_YTD_anterior.mean(axis=1) +0.000000001

            ultimos_3_actual = dataframes[medida].iloc[:, -(recibir_tres_meses):]
            promedio_3_actual = ultimos_3_actual.mean(axis=1) +0.000000001
            ult_3_anteriores = dataframes[medida].iloc[:, -(recibir_ytd + recibir_tres_meses):-(recibir_ytd)]
            promedio_3_anteriores = ult_3_anteriores.mean(axis=1) + 0.000000001

            ultimos_mes_actual = dataframes[medida].iloc[:, -(1):]
            promedio_mes_actual = ultimos_mes_actual.mean(axis=1) + 0.000000001
            ult_mes_anteriores = dataframes[medida].iloc[:, -(recibir_ytd + 1):-(recibir_ytd)]
            promedio_mes_anteriores = ult_mes_anteriores.mean(axis=1) + 0.000000001

            if medida == "Penetración (%)":
                nuevo_dataframe = pd.DataFrame({f"{medida}": primera_columna, 
                                            'Promedio mensual últ 12 mesess': promedio_año_actual, 
                                            'Dif vs PY': (promedio_año_actual - promedio_meses_promedio_PY),
                                            ' ': None,
                                            'Promedio mensual YTD': promedio_YTD_ctual,
                                            ' Dif vs PY': (promedio_YTD_ctual - promedio_YTD_anterior),
                                            '  ': None,
                                            'Promedio mensual últ 3 meses': promedio_3_actual,
                                            '  Dif vs PY': (promedio_3_actual - promedio_3_anteriores),
                                            '   ': None,
                                            'Promedio últ mes': promedio_mes_actual,
                                            '   Dif vs PY': (promedio_mes_actual - promedio_mes_anteriores)                                        
                                            })
            else: 
                nuevo_dataframe = pd.DataFrame({f"{medida}": primera_columna, 
                                            'Promedio mensual últ 12 mesess': promedio_año_actual, 
                                            '%Variacion vs PY': (promedio_año_actual /promedio_meses_promedio_PY) -1,
                                            ' ': None,
                                            'Promedio mensual YTD': promedio_YTD_ctual,
                                            ' %Variacion vs PY': (promedio_YTD_ctual / promedio_YTD_anterior)-1,
                                            '  ': None,
                                            'Promedio mensual últ 3 meses': promedio_3_actual,
                                            '  %Variacion vs PY': (promedio_3_actual / promedio_3_anteriores)-1,
                                            '   ': None,
                                            'Promedio últ mes': promedio_mes_actual,
                                            '   %Variacion vs PY': (promedio_mes_actual / promedio_mes_anteriores) -1                                       
                                            })           
            nuevos_dataframes[medida] = nuevo_dataframe

        return nuevos_dataframes


    def _escribir_en_excel(self, dataframes):
        wb = load_workbook(filename=self.filename)
        ws = wb.create_sheet(self.nombre_hoja)

        last_row = 5
        for medida in self.messures.values():
            rows = dataframe_to_rows(dataframes[medida], index=False, header=True)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=last_row + r_idx, column=c_idx, value=value)

            ws.cell(row=last_row + r_idx + 1, column=1, value="")
            last_row = ws.max_row
        
        for fila in ws.iter_rows():
            if any(celda.value for celda in fila):
                Funciones_Formato.merge_cells_in_column(ws, fila[0].row)

        Funciones_Formato.aplicar_formato(ws, self.nombre_hoja)
        Funciones_Formato.Aplicar_nombres_relleno(ws, self.nombre_hoja)
        wb.save(self.filename)

    def procesar(self):
        datas = self._ordenar_df()
        dato = self._calcular_promedio_ultimos_12_meses(datas)
        self._escribir_en_excel(dato)


