import Funciones_Formato
import pandas as pd
import re
from openpyxl import load_workbook
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import random
import pandas as pd
import re
import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import random

class DatosMarcasPostres:
    def __init__(self, df, filename, nombre_hoja):
        self.df = df
        self.filename = filename
        self.nombre_hoja = nombre_hoja
        self.Messures = {
            "Weighted PENET": "Penetración (%)",
            "Weighted VO1_BUY": "Compra media (kg)",
            "Weighted VO1_DAY": "Compra por acto (kg)",
            "Weighted FREQ": "Frecuencia (veces)"
        }
        self.Indices = {
            "L_REGION = Total Argentina\\Total L_REGION": "Principal",
            "WPOBRPO = T. Danette\\Total WPOBRPO": "T. Danette",
            "WPOBRPO = T. Serenito\\Total WPOBRPO": "T. Serenito"
        }

    def _encontrar_fila_con_fechas(self, datos):
        dfo = self._ordenar_df(datos)
        df = self._seleccionar_dataframe_azar(dfo)
        pattern = r"P3M (Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2}"
        numero_de_fila_con_fechas = None
        for indice, fila in df.iterrows():
            if any(fila.astype(str).str.match(pattern)):
                numero_de_fila_con_fechas = indice
                break

        dfdatos = df.iloc[numero_de_fila_con_fechas, :].tolist()
        dfdatos = [x.strftime('%b-%y') if isinstance(x, datetime.datetime) else x for x in dfdatos]
        return dfdatos

    def _seleccionar_dataframe_azar(self, dataframes):
        medidas = list(dataframes.keys())
        medida_azar = random.choice(medidas)
        return dataframes[medida_azar]

    def _obtener_dataframes_por_indices(self, dataframes, palabras_a_buscar):
        cadenas_presentes = [cadena for cadena in palabras_a_buscar if cadena in dataframes.columns]
        posiciones = {cadena: dataframes.columns.get_loc(cadena) for cadena in cadenas_presentes}
        return dict(sorted(posiciones.items(), key=lambda item: item[1]))

    def _ordenar_df(self, df):
        posiciones = self._obtener_dataframes_por_indices(df, self.Indices)
        dataframes = {}
        medidas = list(posiciones.keys())
        for indice, medida in enumerate(medidas):
            posicion_inicio = posiciones[medida]
            if indice < len(medidas) - 1:
                siguiente_region = medidas[indice + 1]
                siguiente_posicion_inicio = posiciones[siguiente_region]
                posicion_fin = siguiente_posicion_inicio - 1
            else:
                posicion_fin = df.shape[1]

            titulo = self.Indices[medida]
            dataframes[titulo] = df.iloc[:, posicion_inicio:posicion_fin+1]
            dataframes[titulo] = dataframes[titulo].dropna(axis=1, how='all')

        return dataframes

    def _obtener_dataframes_por_titulo_y_region(self, df):
        fechas = self._encontrar_fila_con_fechas(df)

        def _obtener_dataframes_por_titulo_y_region_interno(data, titulo_a_buscar, palabras_a_buscar):
            primera_columna = data[titulo_a_buscar].iloc[:, 0]
            posiciones_regiones = {}
            for palabra in palabras_a_buscar:
                pattern = re.compile(re.escape(palabra), re.IGNORECASE)
                matches = primera_columna.apply(lambda x: bool(pattern.search(str(x))))
                if any(matches):
                    posiciones_regiones[palabra] = matches.idxmax()
                else:
                    posiciones_regiones[palabra] = None

            dataframes_regiones_separados = {}
            regiones = list(posiciones_regiones.keys())
            for indice, region in enumerate(regiones):
                posicion_inicio = posiciones_regiones[region]

                if indice < len(regiones) - 1:
                    siguiente_region = regiones[indice + 1]
                    siguiente_posicion_inicio = posiciones_regiones[siguiente_region]
                    posicion_fin = siguiente_posicion_inicio - 1
                else:
                    posicion_fin = data[titulo_a_buscar].shape[0] - 1

                nombre = self.Messures[region]
                dataframes_regiones_separados[nombre] = data[titulo_a_buscar].iloc[posicion_inicio:posicion_fin + 1]
                dataframes_regiones_separados[nombre] = dataframes_regiones_separados[nombre].dropna(axis=0, how='all')
                dataframes_regiones_separados[nombre].columns = fechas
                dataframes_regiones_separados[nombre] = dataframes_regiones_separados[nombre].dropna()

                filas_con_ceros = (dataframes_regiones_separados[nombre] == 0).sum(axis=1)
                filas_a_eliminar = filas_con_ceros[filas_con_ceros >11].index
                df_filtrado = dataframes_regiones_separados[nombre].drop(filas_a_eliminar)
                dataframes_regiones_separados[nombre] = df_filtrado

            return dataframes_regiones_separados

        data = self._ordenar_df(df)
        dataframes_finales = {}

        for titulo in self.Indices.values():
            palabras_a_buscar = list(self.Messures.keys())
            ingreso_datos = _obtener_dataframes_por_titulo_y_region_interno(data, titulo, palabras_a_buscar)
            dataframes_finales[titulo] = ingreso_datos

        return dataframes_finales

    def _ubicacion_insertar(self, base_de_datos, nombre_buscar, medida):
        ubicaciones = []
        for index, valor in base_de_datos.iloc[:, 0].items():
            if str(valor).strip() == nombre_buscar:
                ubicaciones.append(index)
                break

        if ubicaciones:
            indice_encontrado = ubicaciones[0]
            filas_a_agregar = self.dfa[nombre_buscar][medida]
            filas_a_agregar = pd.DataFrame.from_dict(filas_a_agregar)

            base_de_datos = pd.concat([base_de_datos.iloc[:indice_encontrado + 1], filas_a_agregar, base_de_datos.iloc[indice_encontrado + 1:]], ignore_index=True)
            return base_de_datos
        else:
            return base_de_datos

    def _calcular_promedio_ultimos_12_meses(self, dataframes):
        nuevos_dataframes = {}
        recibir_ytd = 12
        recibir_tres_meses = 3
        recibir_cuatro_meses = 4
        recibir_2_meses = 2

        for medida in self.Messures.values():
            primera_columna = dataframes[medida].iloc[:, 0]

            ultimos_4_actual = dataframes[medida].iloc[:, -(recibir_cuatro_meses):]
            promedio_4_actual = ultimos_4_actual.mean(axis=1) + 0.0000001
            ult_4_anteriores = dataframes[medida].iloc[:, -(recibir_cuatro_meses + recibir_cuatro_meses):-(recibir_cuatro_meses)]
            promedio_4_anteriores = ult_4_anteriores.mean(axis=1) + 0.0000001

            ultimos_2_actual = dataframes[medida].iloc[:, -(recibir_2_meses):]
            promedio_2_actual = ultimos_2_actual.mean(axis=1) + 0.0000001
            ult_2_anteriores = dataframes[medida].iloc[:, -(recibir_cuatro_meses + recibir_2_meses):-(recibir_cuatro_meses + recibir_2_meses - 2)]
            promedio_2_anteriores = ult_2_anteriores.mean(axis=1) + 0.0000001

            ultimos_mes_actual = dataframes[medida].iloc[:, -1:]
            promedio_mes_actual = ultimos_mes_actual.mean(axis=1) + 0.0000001
            ult_mes_anteriores = dataframes[medida].iloc[:, -(recibir_cuatro_meses + 1):-(recibir_cuatro_meses)]
            promedio_mes_anteriores = ult_mes_anteriores.mean(axis=1) + 0.0000001

            if medida == "Penetración (%)":
                nuevo_dataframe = pd.DataFrame({
                    f"{medida}": primera_columna,
                    'P12M (Promedio últ 4 P3M)': promedio_4_actual,
                    'Dif vs PY': (promedio_4_actual - promedio_4_anteriores),
                    '': None,
                    'P6M (Promedio últ 2 P3M)': promedio_2_actual,
                    'Dif vs PY ': (promedio_2_actual - promedio_2_anteriores),
                    '  ': None,
                    'P3M (últ trimestre movil)': promedio_mes_actual,
                    'Dif vs PY  ': (promedio_mes_actual - promedio_mes_anteriores),
                })
            else:
                try:
                    variacion_mes_actual = (promedio_mes_actual / promedio_mes_anteriores) - 1
                except ZeroDivisionError:
                    variacion_mes_actual = 0
                try:
                    variacion_4 = (promedio_4_actual / promedio_4_anteriores) - 1
                except ZeroDivisionError:
                    variacion_4 = 0
                try:
                    variacion_2 = (promedio_2_actual / promedio_2_anteriores) - 1
                except ZeroDivisionError:
                    variacion_2 = 0

                nuevo_dataframe = pd.DataFrame({
                    f"{medida}": primera_columna,
                    'P12M (Promedio últ 4 P3M)': promedio_4_actual,
                    '%Variacion vs PY': variacion_4,
                    '': None,
                    'P6M (Promedio últ 2 P3M)': promedio_2_actual,
                    '%Variacion vs PY ': variacion_2,
                    '  ': None,
                    'P3M (últ trimestre movil)': promedio_mes_actual,
                    '%Variacion vs PY  ': variacion_mes_actual,
                })

            nuevos_dataframes[medida] = nuevo_dataframe

        return nuevos_dataframes

    def _escribir_en_excel(self, dataframes):
        wb = load_workbook(filename=self.filename)
        ws = wb.create_sheet(self.nombre_hoja)

        last_row = 5
        for medida in self.Messures.values():
            rows = dataframe_to_rows(dataframes[medida], index=False, header=True)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=last_row + r_idx, column=c_idx, value=value)

            ws.cell(row=last_row + r_idx + 1, column=1, value="")
            last_row = ws.max_row

        Funciones_Formato.aplicar_formato(ws,self.nombre_hoja)
        Funciones_Formato.Aplicar_nombres_relleno(ws,self.nombre_hoja)
        wb.save(self.filename)


    def procesar(self):
        self.dfa = self._obtener_dataframes_por_titulo_y_region(self.df)
        nueva_base = {}
        lista_primera = list(self.dfa.keys())

        for cambio in list(self.Messures.values()):
            base_de_datos = self.dfa[lista_primera[0]]
            base_de_datos = pd.DataFrame.from_dict(base_de_datos[cambio])
            base_de_datos = base_de_datos.reset_index(drop=True)

            for primera in lista_primera[1:]:
                base = self._ubicacion_insertar(base_de_datos, primera, cambio)
                base = Funciones_Formato.cambio_de_nombre_por_nombre_DE_HOJA(base, self.nombre_hoja, primera)
                base_de_datos = base
            nueva_base[cambio] = base

        dato = self._calcular_promedio_ultimos_12_meses(nueva_base)
        self._escribir_en_excel(dato)



#nombre_excel = "a3.xlsx" ## nombre del archivo final
#directorio_actual = os.path.dirname(os.path.abspath(__file__))
#direccion_base1 = os.path.join(directorio_actual, "Datos_05.xlsx")
#df = pd.read_excel(direccion_base1,sheet_name="Marcas_PO")
#nombre_excel = "Libro1.xlsx"
#processor =CalculoMarcasPostres(df,nombre_excel, "Marcas_PO")
#processor.procesar()