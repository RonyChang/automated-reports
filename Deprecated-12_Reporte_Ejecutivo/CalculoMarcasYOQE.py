import Funciones_Formato
import pandas as pd
import re
from openpyxl import load_workbook
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import random
import pandas as pd
import re
import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import random

class DatosMarcasYOQE:
    def __init__(self, df, filename, nombre_hoja):
        self.df = df
        self.filename = filename
        self.nombre_hoja = nombre_hoja
        self.messures = {
            "Weighted PENET": "Penetración (%)",
            "Weighted VO1_BUY": "Compra media (kg)",
            "Weighted VO1_DAY": "Compra por acto (kg)", 
            "Weighted FREQ": "Frecuencia (veces)"
        }
        self.indices = Funciones_Formato.indices_marca_yogures_postres_queso(self.nombre_hoja)

    def _encontrar_fila_con_fechas(self, datos):
        dfo = self._ordenar_df(datos)
        df = self._seleccionar_dataframe_azar(dfo)
        pattern = r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2}"
        numero_de_fila_con_fechas = None
        for indice, fila in df.iterrows():
            if any(fila.astype(str).str.match(pattern)):
                numero_de_fila_con_fechas = indice
                break

        dfdatos = df.iloc[numero_de_fila_con_fechas, :].tolist()
        dfdatos = [x.strftime('%b-%y') if isinstance(x, datetime.datetime) else x for x in dfdatos]
        return dfdatos

    def _seleccionar_dataframe_azar(self, dataframes):
        medidas = list(dataframes.keys())
        medida_azar = random.choice(medidas)
        dataframe_azar = dataframes[medida_azar]
        return dataframe_azar

    def _obtener_dataframes_por_indices(self, dataframes, palabras_a_buscar):
        cadenas_presentes = [cadena for cadena in palabras_a_buscar if cadena in dataframes.columns]
        posiciones = {cadena: dataframes.columns.get_loc(cadena) for cadena in cadenas_presentes}
        return dict(sorted(posiciones.items(), key=lambda item: item[1]))

    def _ordenar_df(self, df):
        posiciones = self._obtener_dataframes_por_indices(df, self.indices)
        dataframes = {}
        medidas = list(posiciones.keys())
        for indice, medida in enumerate(medidas):
            posicion_inicio = posiciones[medida]
            if indice < len(medidas) - 1:
                siguiente_region = medidas[indice + 1]
                siguiente_posicion_inicio = posiciones[siguiente_region]
                posicion_fin = siguiente_posicion_inicio - 1
            else:
                posicion_fin = df.shape[1]

            titulo = self.indices[medida]
            dataframes[titulo] = df.iloc[:, posicion_inicio:posicion_fin+1]
            dataframes[titulo] = dataframes[titulo].dropna(axis=1, how='all')

        return dataframes

    def _obtener_dataframes_por_titulo_y_region(self):
        fechas = self._encontrar_fila_con_fechas(self.df)

        def _obtener_dataframes_por_titulo_y_region_interno(data, titulo_a_buscar, palabras_a_buscar):
            primera_columna = data[titulo_a_buscar].iloc[:, 0]
            posiciones_regiones = {}
            for palabra in palabras_a_buscar:
                pattern = re.compile(re.escape(palabra), re.IGNORECASE)
                matches = primera_columna.apply(lambda x: bool(pattern.search(str(x))))
                if any(matches):
                    posiciones_regiones[palabra] = matches.idxmax()
                else:
                    posiciones_regiones[palabra] = None

            dataframes_regiones_separados = {}
            regiones = list(posiciones_regiones.keys())
            for indice, region in enumerate(regiones):
                posicion_inicio = posiciones_regiones[region]

                if indice < len(regiones) - 1:
                    siguiente_region = regiones[indice + 1]
                    siguiente_posicion_inicio = posiciones_regiones[siguiente_region]
                    posicion_fin = siguiente_posicion_inicio - 1
                else:
                    posicion_fin = data[titulo_a_buscar].shape[0] - 1

                nombre = self.messures[region]
                dataframes_regiones_separados[nombre] = data[titulo_a_buscar].iloc[posicion_inicio:posicion_fin + 1]
                dataframes_regiones_separados[nombre] = dataframes_regiones_separados[nombre].dropna(axis=0, how='all')
                dataframes_regiones_separados[nombre].columns = fechas
                dataframes_regiones_separados[nombre] = dataframes_regiones_separados[nombre].dropna()

                filas_con_ceros = (dataframes_regiones_separados[nombre] == 0).sum(axis=1)
                filas_a_eliminar = filas_con_ceros[filas_con_ceros > 11].index
                df_filtrado = dataframes_regiones_separados[nombre].drop(filas_a_eliminar)
                dataframes_regiones_separados[nombre] = df_filtrado

            return dataframes_regiones_separados

        data = self._ordenar_df(self.df)
        dataframes_finales = {}

        for titulo in self.indices.values():
            palabras_a_buscar = list(self.messures.keys())
            ingreso_datos = _obtener_dataframes_por_titulo_y_region_interno(data, titulo, palabras_a_buscar)
            dataframes_finales[titulo] = ingreso_datos

        return dataframes_finales

    def _ubicacion_insertar(self, base_de_datos, nombre_buscar, medida, dfa):
        ubicaciones = []
        for index, valor in base_de_datos.iloc[:, 0].items():
            if str(valor).strip() == nombre_buscar:
                ubicaciones.append(index)
                break

        if ubicaciones:
            indice_encontrado = ubicaciones[0]

            filas_a_agregar = dfa[nombre_buscar][medida]
            filas_a_agregar = pd.DataFrame.from_dict(filas_a_agregar)

            base_de_datos = pd.concat([base_de_datos.iloc[:indice_encontrado + 1], filas_a_agregar, base_de_datos.iloc[indice_encontrado + 1:]], ignore_index=True)
            return base_de_datos
        else:
            return base_de_datos

    def _crear_nueva_base(self, dfa):
        nueva_base = {}
        lista_primera = list(dfa.keys())

        for cambio in list(self.messures.values()):
            base_de_datos = dfa[lista_primera[0]]
            base_de_datos = pd.DataFrame.from_dict(base_de_datos[cambio])
            base_de_datos = base_de_datos.reset_index(drop=True)

            for primera in lista_primera[1:]:
                base = self._ubicacion_insertar(base_de_datos, primera, cambio, dfa)
                base = Funciones_Formato.cambio_de_nombre_por_nombre_DE_HOJA(base, self.nombre_hoja, primera)
                base_de_datos = base
            nueva_base[cambio] = base

        return nueva_base

    def _separar_fechas_YTD(self, variable):
        elementos = str(variable).split('-')
        YTD_mes = elementos[0]
        YTD_año = elementos[1]
        numero_año_anterior = int(YTD_año) - 1
        numero_año_antes = str(numero_año_anterior)
        return YTD_mes, YTD_año, numero_año_antes

    def _calcular_promedio_ultimos_12_meses(self, dataframes):
        nuevos_dataframes = {}
        recibir_ytd = 12
        recibir_tres_meses = 3
        for medida in self.messures.values():
            primera_columna = dataframes[medida].iloc[:, 0]
            ultimos_meses_año_actual = dataframes[medida].iloc[:, -recibir_ytd:]
            promedio_año_actual = ultimos_meses_año_actual.mean(axis=1) + 0.0000000000001
            ult_meses_promedio_PY = dataframes[medida].iloc[:, -(recibir_ytd + recibir_ytd):-recibir_ytd]
            promedio_meses_promedio_PY = ult_meses_promedio_PY.mean(axis=1) + 0.0000000000001

            ultima_columna = dataframes[medida].columns[-1]
            YTD_mes, YTD_año, numero_año_antes = self._separar_fechas_YTD(ultima_columna)

            ultimos_YTD_actual = dataframes[medida].loc[:, ('Jan-' + YTD_año):(ultima_columna)]
            promedio_YTD_ctual = ultimos_YTD_actual.mean(axis=1) + 0.0000000000001
            ult_YTD_anterior = dataframes[medida].loc[:, ('Jan-' + numero_año_antes):(YTD_mes + '-' + numero_año_antes)]
            promedio_YTD_anterior = ult_YTD_anterior.mean(axis=1) + 0.0000000000001

            ultimos_3_actual = dataframes[medida].iloc[:, -(recibir_tres_meses):]
            promedio_3_actual = ultimos_3_actual.mean(axis=1) + 0.0000000000001
            ult_3_anteriores = dataframes[medida].iloc[:, -(recibir_ytd + recibir_tres_meses):-(recibir_ytd)]
            promedio_3_anteriores = ult_3_anteriores.mean(axis=1) + 0.0000000000001

            ultimos_mes_actual = dataframes[medida].iloc[:, -(1):]
            promedio_mes_actual = ultimos_mes_actual.mean(axis=1) + 0.0000000000001
            ult_mes_anteriores = dataframes[medida].iloc[:, -(recibir_ytd + 1):-(recibir_ytd)]
            promedio_mes_anteriores = ult_mes_anteriores.mean(axis=1) + 0.0000000000001

            if medida == "Penetración (%)":
                nuevo_dataframe = pd.DataFrame({f"{medida}": primera_columna, 
                                                'Promedio mensual últ 12 mesess': promedio_año_actual, 
                                                'Dif vs PY': (promedio_año_actual - promedio_meses_promedio_PY),
                                                ' ': None,
                                                'Promedio mensual YTD': promedio_YTD_ctual,
                                                ' Dif vs PY': (promedio_YTD_ctual - promedio_YTD_anterior),
                                                '  ': None,
                                                'Promedio mensual últ 3 meses': promedio_3_actual,
                                                '  Dif vs PY': (promedio_3_actual - promedio_3_anteriores),
                                                '   ': None,
                                                'Promedio últ mes': promedio_mes_actual,
                                                '   Dif vs PY': (promedio_mes_actual - promedio_mes_anteriores)                                        
                                                })
            else: 
                nuevo_dataframe = pd.DataFrame({f"{medida}": primera_columna, 
                                                'Promedio mensual últ 12 mesess': promedio_año_actual, 
                                                '%Variacion vs PY': (promedio_año_actual / promedio_meses_promedio_PY) - 1,
                                                ' ': None,
                                                'Promedio mensual YTD': promedio_YTD_ctual,
                                                ' %Variacion vs PY': (promedio_YTD_ctual / promedio_YTD_anterior) - 1,
                                                '  ': None,
                                                'Promedio mensual últ 3 meses': promedio_3_actual,
                                                '  %Variacion vs PY': (promedio_3_actual / promedio_3_anteriores) - 1,
                                                '   ': None,
                                                'Promedio últ mes': promedio_mes_actual,
                                                '   %Variacion vs PY': (promedio_mes_actual / promedio_mes_anteriores) - 1                                       
                                                })                
            nuevos_dataframes[medida] = nuevo_dataframe

        return nuevos_dataframes

    def escribir_en_excel(self, dataframes):
        wb = load_workbook(filename=self.filename)
        ws = wb.create_sheet(self.nombre_hoja)

        last_row = 5
        for medida in self.messures.values():
            rows = dataframe_to_rows(dataframes[medida], index=False, header=True)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=last_row + r_idx, column=c_idx, value=value)

            ws.cell(row=last_row + r_idx + 1, column=1, value="")
            last_row = ws.max_row

        for fila in ws.iter_rows():
            if any(celda.value for celda in fila):
                Funciones_Formato.merge_cells_in_column(ws, fila[0].row)

        Funciones_Formato.aplicar_formato(ws,self.nombre_hoja)
        Funciones_Formato.Aplicar_nombres_relleno(ws,self.nombre_hoja)
        wb.save(self.filename)

    def procesar(self):
        dfa = self._obtener_dataframes_por_titulo_y_region()
        nueva_base = self._crear_nueva_base(dfa)
        dato = self._calcular_promedio_ultimos_12_meses(nueva_base)
        self.escribir_en_excel(dato)

#nombre_excel = "a3.xlsx" ## nombre del archivo final
#directorio_actual = os.path.dirname(os.path.abspath(__file__))
#direccion_base1 = os.path.join(directorio_actual, "Datos_05.xlsx")
#df = pd.read_excel(direccion_base1,sheet_name="Marcas_yog")
#nombre_excel = "Libro1.xlsx"
#processor =MarcasYogProcessor(df,nombre_excel, "Marcas_yog")
#processor.procesar()