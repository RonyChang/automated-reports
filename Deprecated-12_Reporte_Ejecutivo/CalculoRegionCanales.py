import Funciones_Formato
import pandas as pd
import re
from openpyxl import load_workbook
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import random
import pandas as pd
import re
import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import random

class DatosCanalesRegion:
    def __init__(self, df, filename1, nombre_hoja):
        self.df = df
        self.filename1 = filename1
        self.nombre_hoja = nombre_hoja
        self.messures = {
            "Weighted PENET": "Penetración (%)",
            "Weighted VO1_BUY": "Compra media (kg)",
            "Weighted VO1_DAY": "Compra por acto (kg)", 
            "Weighted FREQ": "Frecuencia (veces)"
        }
        self.indices = Funciones_Formato.indices_lectura(nombre_hoja)

    def _encontrar_fila_con_fechas(self, datos):
        dfo = self._ordenar_df(datos)
        df = self._seleccionar_dataframe_azar(dfo)
        pattern = r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2}"
        numero_de_fila_con_fechas = None
        for indice, fila in df.iterrows():
            if any(fila.astype(str).str.match(pattern)):
                numero_de_fila_con_fechas = indice
                break

        dfdatos = df.iloc[numero_de_fila_con_fechas, :].tolist() 
        dfdatos = [x.strftime('%b-%y') if isinstance(x, datetime.datetime) else x for x in dfdatos]
        return dfdatos

    def _seleccionar_dataframe_azar(self, dataframes):
        medidas = list(dataframes.keys())
        medida_azar = random.choice(medidas)
        dataframe_azar = dataframes[medida_azar]
        return dataframe_azar

    def _obtener_dataframes_por_indices(self, dataframes, palabras_a_buscar):
        cadenas_presentes = [cadena for cadena in palabras_a_buscar if cadena in dataframes.columns]
        posiciones = {cadena: dataframes.columns.get_loc(cadena) for cadena in cadenas_presentes}
        return dict(sorted(posiciones.items(), key=lambda item: item[1]))

    def _ordenar_df(self, df):
        posiciones = self._obtener_dataframes_por_indices(df, self.indices)
        dataframes = {}
        medidas = list(posiciones.keys())
        for indice, medida in enumerate(medidas):
            posicion_inicio = posiciones[medida]
            if indice < len(medidas) - 1:
                siguiente_region = medidas[indice + 1]
                siguiente_posicion_inicio = posiciones[siguiente_region]
                posicion_fin = siguiente_posicion_inicio - 1
            else:
                posicion_fin = df.shape[1]

            titulo = self.indices[medida]
            dataframes[titulo] = df.iloc[:, posicion_inicio:posicion_fin+1]
            dataframes[titulo] = dataframes[titulo].dropna(axis=1, how='all')

        return dataframes

    def _obtener_dataframes_por_titulo_y_region(self):
        fechas = self._encontrar_fila_con_fechas(self.df)

        def _obtener_dataframes_por_titulo_y_region_interno(data, titulo_a_buscar, palabras_a_buscar):
            primera_columna = data[titulo_a_buscar].iloc[:, 0]
            posiciones_regiones = {}
            for palabra in palabras_a_buscar:
                pattern = re.compile(re.escape(palabra), re.IGNORECASE)
                matches = primera_columna.apply(lambda x: bool(pattern.search(str(x))))
                if any(matches):
                    posiciones_regiones[palabra] = matches.idxmax()
                else:
                    posiciones_regiones[palabra] = None

            dataframes_regiones_separados = {}
            regiones = list(posiciones_regiones.keys())
            for indice, region in enumerate(regiones):
                posicion_inicio = posiciones_regiones[region]

                if indice < len(regiones) - 1:
                    siguiente_region = regiones[indice + 1]
                    siguiente_posicion_inicio = posiciones_regiones[siguiente_region]
                    posicion_fin = siguiente_posicion_inicio - 1
                else:
                    posicion_fin = data[titulo_a_buscar].shape[0] - 1

                nombre = self.messures[region]
                dataframes_regiones_separados[nombre] = data[titulo_a_buscar].iloc[posicion_inicio:posicion_fin + 1]
                dataframes_regiones_separados[nombre] = dataframes_regiones_separados[nombre].dropna(axis=0, how='all')
                dataframes_regiones_separados[nombre].columns = fechas
                dataframes_regiones_separados[nombre] = dataframes_regiones_separados[nombre].dropna()

            return dataframes_regiones_separados

        data = self._ordenar_df(self.df)
        dataframes_finales = {}

        for titulo in self.indices.values():
            palabras_a_buscar = list(self.messures.keys())
            ingreso_datos = _obtener_dataframes_por_titulo_y_region_interno(data, titulo, palabras_a_buscar)
            dataframes_finales[titulo] = ingreso_datos

        return dataframes_finales

    def _cacular_promedio_geograficamente(self, dataframe):
        nuevos_dataframes = {}
        recibir_ytd = 12
        recibir_tres_meses = 3

        for indicador in self.indices.values():
            new_dataframe = {}

            for medida in self.messures.values():
                primera_columna = dataframe[indicador][medida].iloc[:, 0]
                ultimos_meses_año_actual = dataframe[indicador][medida].iloc[:, -recibir_ytd:]
                promedio_año_actual = ultimos_meses_año_actual.mean(axis=1) + 0.0000000000001
                ult_meses_promedio_PY = dataframe[indicador][medida].iloc[:, -(recibir_ytd + recibir_ytd):-recibir_ytd]
                promedio_meses_promedio_PY = ult_meses_promedio_PY.mean(axis=1) + 0.0000000000001

                ultimos_3_actual = dataframe[indicador][medida].iloc[:, -(recibir_tres_meses):]
                promedio_3_actual = ultimos_3_actual.mean(axis=1) + 0.0000000000001
                ult_3_anteriores = dataframe[indicador][medida].iloc[:, -(recibir_ytd + recibir_tres_meses):-(recibir_ytd)]
                promedio_3_anteriores = ult_3_anteriores.mean(axis=1) + 0.0000000000001

                ultimos_mes_actual = dataframe[indicador][medida].iloc[:, -1:]
                promedio_mes_actual = ultimos_mes_actual.mean(axis=1) + 0.0000000000001
                ult_mes_anteriores = dataframe[indicador][medida].iloc[:, -(recibir_ytd + 1):-(recibir_ytd)]
                promedio_mes_anteriores = ult_mes_anteriores.mean(axis=1) + 0.0000000000001

                if medida == "Penetración (%)" or medida == "% Volumen":
                    nuevo_dataframe = pd.DataFrame({
                        ('', f"{medida}"): primera_columna,
                        (f"{indicador}", 'Promedio mensual últ 12 meses'): promedio_año_actual,
                        (f"{indicador} ", 'Dif vs PY'): (promedio_año_actual - promedio_meses_promedio_PY),
                        (f"{indicador}  ", ''): None,
                        (f"{indicador}   ", 'Promedio mensual últ 3 meses'): promedio_3_actual,
                        (f"{indicador}    ", 'Dif vs PY '): (promedio_3_actual - promedio_3_anteriores),
                        (f"{indicador}  ", '  '): None,
                        (f"{indicador}   ", 'Promedio últ mes'): promedio_mes_actual,
                        (f"{indicador}    ", 'Dif vs PY  '): (promedio_mes_actual - promedio_mes_anteriores) 
                    })
                else:
                    try:
                        variacion_año_actual = promedio_año_actual / promedio_meses_promedio_PY - 1
                    except ZeroDivisionError:
                        variacion_año_actual = 0

                    try:
                        variacion_3_meses = promedio_3_actual / promedio_3_anteriores - 1
                    except ZeroDivisionError:
                        variacion_3_meses = 0

                    try:
                        variacion_mes_actual = promedio_mes_actual / promedio_mes_anteriores - 1
                    except ZeroDivisionError:
                        variacion_mes_actual = 0

                    nuevo_dataframe = pd.DataFrame({
                        f"{medida}": primera_columna,
                        'Promedio mensual últ 12 meses': promedio_año_actual,
                        '%Variacion vs PY': variacion_año_actual,
                        '': None,
                        'Promedio mensual últ 3 meses': promedio_3_actual,
                        '%Variacion vs PY ': variacion_3_meses,
                        '   ': None,
                        'Promedio últ mes': promedio_mes_actual,
                        '   %Variacion vs PY': variacion_mes_actual
                    })

                new_dataframe[medida] = nuevo_dataframe
            nuevos_dataframes[indicador] = new_dataframe

        return nuevos_dataframes

    def _escribir_en_excel(self, dataframes):
        wb = load_workbook(filename=self.filename1)
        ws = wb.create_sheet(self.nombre_hoja)

        max_col = 0
        for indicador in self.indices.values():
            max_col = max_col + 0  # Reiniciar la columna máxima para cada indicador
            last_row = 5  # Empieza desde la fila 6
            for medida in self.messures.values():
                rows = dataframe_to_rows(dataframes[indicador][medida], index=False, header=True)
                for r_idx, row in enumerate(rows, 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=last_row + r_idx, column=c_idx + max_col, value=value)

                last_row = last_row + r_idx + 1

            ws.cell(row=1, column=ws.max_column + 1, value="")
            max_col = ws.max_column

        for fila in ws.iter_rows(min_row=1, max_row=6):
            if any(celda.value for celda in fila):
                Funciones_Formato.merge_cells_in_column(ws, fila[0].row)
        Funciones_Formato.aplicar_formato_region_segmento(ws,self.nombre_hoja)
        celdas_combinadas = Funciones_Formato.encontrar_celdas_combinadas_en_filas(ws, range(1, 11))
        Funciones_Formato.aplicar_relleno_y_color_a_celdas_combinadas(ws, celdas_combinadas)
        Funciones_Formato.aplicar_tamaño_regiones_canales(ws, celdas_combinadas)
        Funciones_Formato.Aplicar_nombres_relleno(ws,self.nombre_hoja)
        wb.save(self.filename1)

    def procesar(self):
        diccionario_datos = self._obtener_dataframes_por_titulo_y_region()
        BD = self._cacular_promedio_geograficamente(diccionario_datos)
        self._escribir_en_excel(BD)

#nombre_excel = "a3.xlsx" ## nombre del archivo final
#directorio_actual = os.path.dirname(os.path.abspath(__file__))
#direccion_base1 = os.path.join(directorio_actual, "Datos_05.xlsx")
#df = pd.read_excel(direccion_base1,sheet_name="Regiones_yog")
#nombre_excel = "Libro1.xlsx"
#processor =CanalesRegionProcessor(df,nombre_excel, "Regiones_yog")
#processor.procesar()