import Funciones_Formato
import pandas as pd
import re
from openpyxl import load_workbook
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import random
import pandas as pd
import re
import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import random

class DatosSegmentoPostre:
    def __init__(self, df, filename, nombre_hoja):
        self.df = df
        self.filename = filename
        self.nombre_hoja = nombre_hoja
        self.messures = {
            "Weighted PENET": "Penetración (%)", 
            "Weighted VO1_BUY": "Compra media (kg)",
            "Weighted VO1_DAY": "Compra por acto (kg)", 
            "Weighted FREQ": "Frecuencia (veces)"
        }

    def _obtener_dataframes_por_titulo_y_region(self, dataframes):
        primera_columna = dataframes.iloc[:, 0]
        posiciones_regiones = {}
        for palabra in self.messures.keys():
            pattern = re.compile(re.escape(palabra), re.IGNORECASE)
            matches = primera_columna.apply(lambda x: bool(pattern.search(str(x))))
            if any(matches):
                posiciones_regiones[palabra] = matches.idxmax()
            else:
                posiciones_regiones[palabra] = None
        return dict(sorted(posiciones_regiones.items(), key=lambda item: item[1]))

    def _ordenar_df(self):
        posiciones = self._obtener_dataframes_por_titulo_y_region(self.df)
        dataframes = {}
        medidas = list(posiciones.keys())
        for indice, medida in enumerate(medidas):
            posicion_inicio = posiciones[medida]
            if indice < len(medidas) - 1:
                siguiente_region = medidas[indice + 1]
                siguiente_posicion_inicio = posiciones[siguiente_region]
                posicion_fin = siguiente_posicion_inicio - 1
            else:
                posicion_fin = self.df.shape[0] - 1
            titulo = self.messures[medida]
            dataframes[titulo] = self.df.iloc[posicion_inicio:posicion_fin + 1]
            dataframes[titulo] = dataframes[titulo].dropna(axis=0, how='all')
        
        nuevos_dataframes = {}
        for clave, df in dataframes.items():
            df_sin_nan = df.dropna()
            filas_con_ceros = (df_sin_nan == 0).sum(axis=1)
            filas_a_eliminar = filas_con_ceros[filas_con_ceros > 11].index
            df_filtrado = df_sin_nan.drop(filas_a_eliminar)
            nuevos_dataframes[clave] = df_filtrado 
        return nuevos_dataframes

    def _calcular_promedio_ultimos_12_meses(self, dataframes):
        nuevos_dataframes = {}
        recibir_ytd = 12
        recibir_tres_meses = 3
        recibir_cuadtro_meses = 4
        recibir_2_meses = 2

        for medida in self.messures.values(): 
            primera_columna = dataframes[medida].iloc[:, 0]

            ultimos_4_actual = dataframes[medida].iloc[:, -(recibir_cuadtro_meses):]
            promedio_4_actual = ultimos_4_actual.mean(axis=1) + 0.0000001
            ult_4_anteriores = dataframes[medida].iloc[:, -(recibir_cuadtro_meses + recibir_cuadtro_meses):-(recibir_cuadtro_meses)]
            promedio_4_anteriores = ult_4_anteriores.mean(axis=1) + 0.0000001

            ultimos_2_actual = dataframes[medida].iloc[:, -(recibir_2_meses):]
            promedio_2_actual = ultimos_2_actual.mean(axis=1) + 0.0000001
            ult_2_anteriores = dataframes[medida].iloc[:, -(recibir_cuadtro_meses + recibir_2_meses):-(recibir_cuadtro_meses + recibir_2_meses - 2)]
            promedio_2_anteriores = ult_2_anteriores.mean(axis=1) + 0.0000001

            ultimos_mes_actual = dataframes[medida].iloc[:, -(1):]
            promedio_mes_actual = ultimos_mes_actual.mean(axis=1) + 0.0000001
            ult_mes_anteriores = dataframes[medida].iloc[:, -(recibir_cuadtro_meses + 1):-(recibir_cuadtro_meses)]
            promedio_mes_anteriores = ult_mes_anteriores.mean(axis=1) + 0.0000001

            if medida == "Penetración (%)":  
                nuevo_dataframe = pd.DataFrame({f"{medida}": primera_columna,
                        'P12M (Promedio últ 4 P3M)': promedio_4_actual,
                        'Dif vs PY': (promedio_4_actual - promedio_4_anteriores),
                        '': None,
                        'P6M (Promedio últ 2 P3M)': promedio_2_actual,
                        'Dif vs PY ': (promedio_2_actual - promedio_2_anteriores),
                        '  ': None,                           
                        'P3M (últ trimestre movil)': promedio_mes_actual,
                        'Dif vs PY  ': (promedio_mes_actual - promedio_mes_anteriores),
                        })
            else: 
                try:
                    variacion_mes_actual = (promedio_mes_actual / promedio_mes_anteriores) - 1
                except ZeroDivisionError:
                    variacion_mes_actual = 0
                try:
                    variacion_4 = (promedio_4_actual / promedio_4_anteriores) - 1
                except ZeroDivisionError:
                    variacion_4 = 0
                try:
                    variacion_2 = (promedio_2_actual / promedio_2_anteriores) - 1
                except ZeroDivisionError:
                    variacion_2 = 0

                nuevo_dataframe = pd.DataFrame({f"{medida}": primera_columna,
                        'P12M (Promedio últ 4 P3M)': promedio_4_actual,
                        '%Variacion vs PY': variacion_4,
                        '': None,
                        'P6M (Promedio últ 2 P3M)': promedio_2_actual,
                        '%Variacion vs PY ': variacion_2,
                        '  ': None,                           
                        'P3M (últ trimestre movil)': promedio_mes_actual,
                        '%Variacion vs PY  ': variacion_mes_actual,
                        })
                
            nuevos_dataframes[medida] = nuevo_dataframe

        return nuevos_dataframes

    def _escribir_en_excel(self, dataframes):
        wb = load_workbook(filename=self.filename)
        ws = wb.create_sheet(self.nombre_hoja)

        last_row = 5
        for medida in self.messures.values():
            rows = dataframe_to_rows(dataframes[medida], index=False, header=True)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=last_row + r_idx, column=c_idx, value=value)

            ws.cell(row=last_row + r_idx + 1, column=1, value="")
            last_row = ws.max_row

        for fila in ws.iter_rows():
            if any(celda.value for celda in fila):
                Funciones_Formato.merge_cells_in_column(ws, fila[0].row)

        Funciones_Formato.aplicar_formato(ws,self.nombre_hoja)
        Funciones_Formato.Aplicar_nombres_relleno(ws,self.nombre_hoja)
        wb.save(self.filename)

    def procesar(self):
        datas = self._ordenar_df()
        dato = self._calcular_promedio_ultimos_12_meses(datas)
        self._escribir_en_excel(dato)


#nombre_excel = "a3.xlsx" ## nombre del archivo final
#directorio_actual = os.path.dirname(os.path.abspath(__file__))
#direccion_base1 = os.path.join(directorio_actual, "Datos_05.xlsx")
#df = pd.read_excel(direccion_base1,sheet_name="Seg_PO")
#nombre_excel = "Libro1.xlsx"
#processor =CalculoSegmentoPostre(df,nombre_excel, "Seg_PO")
#processor.procesar()