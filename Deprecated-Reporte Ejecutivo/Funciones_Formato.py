import pandas as pd
import random
import re
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment,Border,Side,NamedStyle
from openpyxl.drawing.image import Image
import os
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from pandas import DataFrame
import Funcion_Fecha

fecha_formateada = Funcion_Fecha.obtener_fecha_formateada()
##SUBFUNCIONES
def merge_cells_in_column(hoja, column_letter):
    celdas_por_nombre = {}
    nombre_anterior = None
    i = 0

    for celda in hoja[column_letter]:
        nombre_actual = celda.value
        if isinstance(nombre_actual, str):  # Verificar si el valor es una cadena (palabra)
            nombre_actual = nombre_actual.strip()  # Eliminar espacios en blanco adicionales
            if nombre_actual.isalpha():  # Verificar si es una palabra (sin números ni caracteres especiales)
                if nombre_actual == nombre_anterior and nombre_actual is not None:
                    # Si es el mismo nombre, simplemente agregamos la coordenada al conjunto existente
                    celdas_por_nombre[f"{nombre_actual}_{i}"].append(celda.coordinate)
                else:
                    # Si es un nuevo nombre, creamos un nuevo conjunto de coordenadas
                    i += 1
                    celdas_por_nombre[f"{nombre_actual}_{i}"] = [celda.coordinate]
                nombre_anterior = nombre_actual

    for nombre, celdas in celdas_por_nombre.items():
        if len(celdas) > 1:
            primer_celda = celdas[0]
            ultima_celda = celdas[-1]
            rango_celdas = f"{primer_celda}:{ultima_celda}"
            hoja.merge_cells(rango_celdas)
    
def merge_cells_in_column_geo(hoja, column_letter):
    celdas_por_nombre = {}
    nombre_anterior = None
    i = 0


    for celda in hoja[column_letter]:
        nombre_actual = celda.value
        if isinstance(nombre_actual, str):  # Verificar si el valor es una cadena (palabra)
            nombre_actual = nombre_actual.strip()  # Eliminar espacios en blanco adicionales
            if nombre_actual.isalpha():  # Verificar si es una palabra (sin números ni caracteres especiales)
                if nombre_actual == nombre_anterior and nombre_actual is not None:
                    # Si es el mismo nombre, simplemente agregamos la coordenada al conjunto existente
                    celdas_por_nombre[f"{nombre_actual}_{i}"].append(celda.coordinate)
                else:
                    # Si es un nuevo nombre, creamos un nuevo conjunto de coordenadas
                    i += 1
                    celdas_por_nombre[f"{nombre_actual}_{i}"] = [celda.coordinate]
                nombre_anterior = nombre_actual

    for nombre, celdas in celdas_por_nombre.items():
        if len(celdas) > 1:
            primer_celda = celdas[0]
            ultima_celda = celdas[-1]
            rango_celdas = f"{primer_celda}:{ultima_celda}"
            hoja.merge_cells(rango_celdas)

def indices_lectura(nombre):
    Indices = {
    "WPOBRYG = T. Yogur + Transito + Salud\Total WPOBRYG":"Industria",		
    "WPOBRYG = T. Danone\T. Yogur + Transito + Salud\Total WPOBRYG":"Danone",			
    "WPOBRYG = T. Yogurisimo\T. Danone\T. Yogur + Transito + Salud\Total WPOBRYG":"Yogurisimo",	
    "WPOBRYG = LS Clasico (Danone)\T. Danone\T. Yogur + Transito + Salud\Total WPOBRYG":"LSC"																							
    }
    Indices2 =  {
    "WPOBRQE = T.Quesos Blancos + Fundidos\Total WPOBRQE":"Industria",		
    "WPOBRQE = T. Danone\T.Quesos Blancos + Fundidos\Total WPOBRQE":"Danone",
    "WPOBRQE = Casancrem\T. Danone\T.Quesos Blancos + Fundidos\Total WPOBRQE":"Casancrem",		
    "WPOBRQE = LS Clásico\T. Danone\T.Quesos Blancos + Fundidos\Total WPOBRQE":"LSC",																							
    }


    subcadenas = nombre.split("_") 
    if subcadenas[1]=="yog": 
        return Indices
    else: 
        return Indices2

def encontrar_celdas_combinadas_en_filas(hoja, filas):
    celdas_combinadas = set()

    for fila in filas:
        for celda in hoja[fila]:
            for rango_combinado in hoja.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(rango_combinado.coord)
                if min_row <= celda.row <= max_row and min_col <= celda.column <= max_col:
                    celdas_combinadas.add(celda.coordinate)

    return celdas_combinadas

def aplicar_relleno_y_color_a_celdas_combinadas(hoja, celdas_combinadas):
    fill_rojo = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Relleno rojo

    colores_relleno = {
        'Industria': 'ED7D31',  # Rojo
        'Danone': '5B9BD5',  # Amarillo
        'Yogurisimo': '305496',  # Naranja
        'LSC': 'C00000',  # Morado
        'Casancrem': '92D050',
        'Danette': 'B4C6E7',
        'Serenito': 'C65911',
        'SER': '70AD47'
        }

    # Crear el formato para el borde negro
    borde_blanco = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
        )
    
    # Crear el formato para el borde negro
    borde_negro = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))

    for celda_coord in celdas_combinadas:
        celda = hoja[celda_coord]
        valor_celda = celda.value
        if valor_celda is not None:  # Verificar si la celda tiene un valor
            valor_celda = valor_celda.strip()  # Obtener el valor de la celda y eliminar espacios en blanco
            if valor_celda in colores_relleno:
                color_hex = colores_relleno[valor_celda]
                fill_color = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                celda.fill = fill_color
                celda.border = borde_negro
                celda.font = Font(color='FFFFFF', size=8, name='Arial', bold=True)
                celda.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            else:
                celda.fill = fill_rojo  # Si no hay color definido, aplicar relleno rojo por defecto
                celda.border = borde_negro
                celda.font = Font(color='FFFFFF', size=8, name='Arial', bold=True)
                celda.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

def cambio_de_nombre_por_nombre_DE_HOJA(dataframe,nombre_hoja,primera):
    
    if 'yog' in nombre_hoja: 
        if 'Yogurisimo' in primera: 
            dataframe.iloc[:, 0] = dataframe.iloc[:, 0].replace('    T. Yogur Bebible  ', '     Yogurisimo Bebible  ')
            dataframe.iloc[:, 0] = dataframe.iloc[:, 0].replace('    T. Yogur Cuchareable  ', '     Yogurisimo Cuchareable  ')
        if 'Ser' in primera: 
            dataframe.iloc[:, 0] = dataframe.iloc[:, 0].replace('    T. Yogur Bebible  ', '     Ser Bebible  ')
            dataframe.iloc[:, 0] = dataframe.iloc[:, 0].replace('    T. Yogur Cuchareable  ', '     Ser Cuchareable  ')
        if 'LS Clasico' in primera: 
            dataframe.iloc[:, 0] = dataframe.iloc[:, 0].replace('    T. Yogur Bebible  ', '     LSC Bebible  ')
            dataframe.iloc[:, 0] = dataframe.iloc[:, 0].replace('    T. Yogur Cuchareable  ', '     LSC Cuchareable  ') 
    elif 'PO' in nombre_hoja: 
        if 'Danette' in primera: 
            dataframe.iloc[:, 0] = dataframe.iloc[:, 0].replace('   Flanes  ', '    Flanes  ')
        if 'Serenito' in primera: 
            dataframe.iloc[:, 0] = dataframe.iloc[:, 0].replace('   Postres  ', '     Postres Serenito   ')
            dataframe.iloc[:, 0] = dataframe.iloc[:, 0].replace('   Flanes  ', '     Serenito Flan  ')
    return dataframe
	
#nombres sombreados
def lista_de_sombreoc(categoria):
    if categoria == "Marcas_yog":
        solo_negrita = ["No hay"
        ] 
        solo_azul_sin_negrita =[
            "      Yss Batido Natural  ",
            "      Yogurisimo Cremix  ",
            "      Yogurisimo Griego  ",
            "      Resto Yogurisimo  ",
            "      Ser Calci+  ",
            "      Ser Yogur  "
        ]

        solo_negrita_azul =[
            "    T. Danone  ",
            "     T. Yogurisimo  ",
            "     T. Ser  ",
            "     Gran Compra  ",
            "     LS Clasico (Danone)  ",
            "     Danonino  "
        ]
        negrita_fondo_gris = [ "  T. Yogur + Transito + Salud  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    elif categoria == "Seg_yog":
        solo_negrita = ["  No hay",
        ] 
        solo_azul_sin_negrita =[
            "    T. Yogur Entero  ",
            "    T. Yogur Descremado  ",
            "    T. Cuchareable  ",
            "    T. Bebible  "
        ]
        solo_negrita_azul =[
            "No hay"
        ]
        negrita_fondo_gris = [ "  T. Yogur + Transito + Salud  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    elif categoria == "Marcas_QE":
        solo_negrita = ["no hay"] 
        solo_azul_sin_negrita =[
            "no hay"]
        solo_negrita_azul =[
            "   T. Danone  ","    Casancrem  ","    LS Clásico  "
        ]
        negrita_fondo_gris = [ "  T.Quesos Blancos + Fundidos  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    elif categoria == "Seg_QE":
        solo_negrita = ["no hay"] 
        solo_azul_sin_negrita =[
            "     Quesos Blancos  ","     Quesos Fundidos  "]
        solo_negrita_azul =[
            "no hay"
        ]
        negrita_fondo_gris = [ "    T. Quesos Blancos + Fundidos  ","  T. Quesos Blancos + Fundidos  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    elif categoria == "Marcas_PO":
        solo_negrita = ["no hay"  ]
        solo_azul_sin_negrita =[
            "      Chocolate  ","      Dulce De Leche  ","      Resto Sabores Danette  ",
            "     Serenito Postres  ","      Serenito Simple  "] ##PREGUNTAR FRANCISCO 
        solo_negrita_azul =[
            "   T. Danone  ","    T. Danette  ","     Danette Postre  ","     Danette Flan  ",
            "    T. Ser  ","    T. Serenito  ","    LS Clasico  "
        ]
        negrita_fondo_gris = [ "  T. Pos + Fla + Gel + Azl  ","  T. Pos + Fla + Gel + Azl  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    elif categoria == "Seg_PO":
        solo_negrita = ["no hay"] 
        solo_azul_sin_negrita =[
            "   Postres  ","   Flanes  ","   Gelatinas/Jaleas  ","   Arroz con leche (Azl)  "]
        solo_negrita_azul =[
            "no hay"
        ]
        negrita_fondo_gris = [ "  T. Pos + Fla + Gel + Azl  ","  T. Pos + Fla + Gel + Azl  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    ##REGIONES
    elif categoria == "Regiones_yog":
        solo_negrita = [
        "No  hay",
        ] 
        solo_azul_sin_negrita =[
            "   CABA + GBA (24 PARTIDOS)  ",
            "   Interior  ",
        ]
        solo_negrita_azul =[
            "No hay"
        ]
        negrita_fondo_gris = [ "  Total Argentina  ","  Total L_REGION  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    elif categoria == "Regiones_QE":
        solo_negrita = [
        "No  hay",
        ] 
        solo_azul_sin_negrita =[
            "   CABA + GBA (24 PARTIDOS)  ",
            "   Interior  ",
        ]
        solo_negrita_azul =[
            "No hay"
        ]
        negrita_fondo_gris = [ "  Total Argentina  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    elif categoria == "Regiones_PO":
        solo_negrita = [
        "No  hay",
        ] 
        solo_azul_sin_negrita =[
            "   CABA + GBA (24 PARTIDOS)  ",
            "   Interior  ",
        ]
        solo_negrita_azul =[
            "No hay"
        ]
        negrita_fondo_gris = [ "  Total Argentina  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    ##CANALES
    elif categoria == "Canales_yog":
        solo_negrita = ["No Hay"
        ] 
        solo_azul_sin_negrita =[
            "   Total UTT  ",
            "   Total DTT  ",
        ]
        solo_negrita_azul =[
            "No hay"
        ]
        negrita_fondo_gris = [ "  Total Mercado  ","  Total Market  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    elif categoria == "Canales_QE":
        solo_negrita = ["No Hay"
        ] 
        solo_azul_sin_negrita =[
            "   Total UTT  ",
            "   Total DTT  ",
        ]
        solo_negrita_azul =[
            "No hay"
        ]
        negrita_fondo_gris = [ "  Total Mercado  ","  Total Market  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    elif categoria == "Canales_PO":
        solo_negrita = ["No Hay"
        ] 
        solo_azul_sin_negrita =[
            "   Total UTT  ",
            "   Total DTT  ",
        ]
        solo_negrita_azul =[
            "No hay"
        ]
        negrita_fondo_gris = [ "  Total Mercado  ","  Total Market  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    ##DEMOGRAFICOS
    elif categoria == "Demo_yog":
        solo_negrita = ["No Hay"] 
        solo_azul_sin_negrita =[
            "No hay"]
        solo_negrita_azul =[
            "No hay" ]
        negrita_fondo_gris = [ "  NSE 1  ","  TAMAÑO FAMILIA 1  ","  LIFESTAGE  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    elif categoria == "Demo_QE":
        solo_negrita = ["No Hay"] 
        solo_azul_sin_negrita =[
            "No hay"]
        solo_negrita_azul =[
            "No hay" ]
        negrita_fondo_gris = [ "  NSE 1  ","  TAMAÑO FAMILIA 1  ","  LIFESTAGE  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    elif categoria == "Demo_PO":
        solo_negrita = ["No Hay"] 
        solo_azul_sin_negrita =[
            "No hay"]
        solo_negrita_azul =[
            "No hay" ]
        negrita_fondo_gris = [ "  NSE 1  ","  TAMAÑO FAMILIA 1  ","  LIFESTAGE  "]
        return solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris
    else:
        a=["NSE 1", "TAMAÑO FAMILIA 1", "LIFESTAGE"]
        return  a # Otra categoría no reconocida

#letras
    
def agrupar_secciones_por_palabras_clave(hoja, palabras_clave):

    ##FUENTE A LAS CELDAS
    num_filas = hoja.max_row
    num_columnas = hoja.max_column

    ##validar 
    lista_de_palabras = ["% Volumen", "Penetración (%)","Compra media (kg)","Compra por acto (kg)","Frecuencia (veces)"]
    numeros_filas_activas = []
    
    # Iterar sobre todas las filas de la hoja
    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=1):
        celda = fila[0]  # Obtener la celda de la primera columna en la fila
        palabra = celda.value  # Obtener el valor de la celda
        # Verificar si la celda contiene una palabra que te interesa
        if palabra in lista_de_palabras:
            numero_fila = celda.row  # Obtener el número de la fila
            numeros_filas_activas.append(numero_fila)  # Agregar el número de fila a la lista

    fila_inicial = numeros_filas_activas[0]  # Por ejemplo, si el texto "Dif vs PY" está en la fila 2
    ##validar



    indices_secciones = []

    inicio_seccion = None
    for fila in hoja.iter_rows(min_row=fila_inicial, max_row=num_filas, min_col=1, max_col=1):  # Iterar sobre las filas de la primera columna
        texto_celda = fila[0].value  # Obtener el texto de la celda en la primera columna de la fila
        if texto_celda is not None:
            fila_sin_espacios = texto_celda.strip()  # Eliminar espacios en blanco
            if fila_sin_espacios in palabras_clave:
                if inicio_seccion is None:
                    inicio_seccion = fila[0].row
                else:
                    fin_seccion = fila[0].row 
                    indices_secciones.append((inicio_seccion, fin_seccion))
                    inicio_seccion = None

    pares_indices = [(indices_secciones[i], indices_secciones[i + 1]) for i in range(0, len(indices_secciones), 2)]

    # Imprimir la lista de pares de índices
    #print(pares_indices)

    # Agregar el último grupo si es necesario
    if inicio_seccion is not None:
        fin_seccion = hoja.max_row
        indices_secciones.append((inicio_seccion, fin_seccion))

    for inicio, fin in indices_secciones:
        hoja.row_dimensions.group(inicio, fin, hidden=True)

##NUEVO PO 

def Aplicar_nombres_relleno(hoja,nombre_hoja):
    nombre_del_mes = fecha_formateada
    celda_inicio=1
    fuente_negrita = Font(color='000000', name='Arial', size=8, bold=True)
    fuente_roja =  Font(color='FF0000', name='Arial', size=8)

    ##APLICAR COLOR A LAS CELDAS 
    if nombre_hoja == "Marcas_yog":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Yogurt")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Canales - T. Argentina")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
        lista1= ["Yss Batido Natural","Resto Yogurisimo"]
        #lista2= ["Ser Calci+","Ser Calci+"] #28/04
        lista3= ["Sancor Yogs","Sancor Vida"]
        lista4= ["Día%","Resto MDD"]
        agrupar_secciones_por_palabras_clave(hoja,lista1)
        #agrupar_secciones_por_palabras_clave(hoja,lista2)
        agrupar_secciones_por_palabras_clave(hoja,lista3)
        agrupar_secciones_por_palabras_clave(hoja,lista4)
        
        
    elif nombre_hoja == "Seg_yog":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Yogurt")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Canales - T. Argentina - T. Industria")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
        lista1= ["Entero Sachet","Entero Botellita"]
        lista2= ["Entero Firme","Entero Batido"]
        lista3= ["Entero Toppings-Cereales","Entero sin Toppings"]
        lista4= ["Descremado Sachet","Descremado Botellita"] 
        lista5= ["Descremado Firme","Descremado Batido"] 
        lista6= ["Des. Toppings-Cereales","Des. sin Toppings"]        
        agrupar_secciones_por_palabras_clave(hoja,lista1)
        agrupar_secciones_por_palabras_clave(hoja,lista2)
        agrupar_secciones_por_palabras_clave(hoja,lista3)
        agrupar_secciones_por_palabras_clave(hoja,lista4)
        agrupar_secciones_por_palabras_clave(hoja,lista5)
        agrupar_secciones_por_palabras_clave(hoja,lista6)


    elif nombre_hoja == "Regiones_yog":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Yogurt")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Canales  - T. Industria")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
    elif nombre_hoja == "Canales_yog":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Yogurt")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Argentina - T. Industria")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
        lista1= ["Total Carrefour","Otras Cadenas"]        
        agrupar_secciones_por_palabras_clave(hoja,lista1)


    elif nombre_hoja == "Demo_yog":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Yogurt")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Canales - T. Argentina")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
    ##QUESO
    elif nombre_hoja == "Marcas_QE":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Queso Untable")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Canales - T. Argentina")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
    elif nombre_hoja == "Seg_QE":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Queso Untable")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Canales - T. Argentina - T. Industria")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
        lista1= ["Blancos Regulares Plain","Blancos Regulares Sabor"]
        lista3= ["Fundidos Regulares Plain","Fundidos Regulares Sabor"]     
        agrupar_secciones_por_palabras_clave(hoja,lista1)
        agrupar_secciones_por_palabras_clave(hoja,lista3)       

    elif nombre_hoja == "Regiones_QE":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Queso Untable")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Canales  - T. Industria")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
    elif nombre_hoja == "Canales_QE":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Queso Untable")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Argentina - T. Industria")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
        lista1= ["Total Carrefour","Other Retailers"]  
        agrupar_secciones_por_palabras_clave(hoja,lista1)
   
    elif nombre_hoja == "Demo_QE":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Queso Untable")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Canales - T. Argentina")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
    ##POSTRES
    elif nombre_hoja == "Marcas_PO":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Postres Regriferados")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Canales - T. Argentina")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
        
        lista2= ["Danette Postre","Danette Flan"] 
        lista3= ["Serenito Postres","Serenito Simple"] ##serenito postres no esta 30/07/2024
        lista1= ["Chocolate","Resto Sabores Danette"]     

        agrupar_secciones_por_palabras_clave(hoja,lista2)
        #agrupar_secciones_por_palabras_clave(hoja,lista3) 


    elif nombre_hoja == "Seg_PO":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Postres Regriferados")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Canales - T. Argentina - T. Industria")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
        lista1= ["Postres Adultos","Postres Infantiles"]
        lista2= ["Flanes Adultos","Flanes Infantiles"] 
        lista3= ["Gelatinas/Jaleas Adultos","Gelatinas/Jaleas Infantiles"]
        lista4= ["Pote Simple","Pote Toppings"]     
        agrupar_secciones_por_palabras_clave(hoja,lista1)
        agrupar_secciones_por_palabras_clave(hoja,lista2)
        agrupar_secciones_por_palabras_clave(hoja,lista3)
        agrupar_secciones_por_palabras_clave(hoja,lista4) 

    elif nombre_hoja == "Regiones_PO":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Postres Regriferados")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Canales  - T. Industria")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
    elif nombre_hoja == "Canales_PO":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Postres Regriferados")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Argentina - T. Industria")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
        lista1= ["Total Carrefour","Otras Cadenas"]  
        agrupar_secciones_por_palabras_clave(hoja,lista1)
    elif nombre_hoja == "Demo_PO":
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Postres Regriferados")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Canales - T. Argentina")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")
    else:
        celda_inferior1 = hoja.cell(row=celda_inicio+1, column=1, value="Categoría: Postres Regriferados")
        celda_inferior2 = hoja.cell(row=celda_inicio+2, column=1, value=f"Mes: {nombre_del_mes}")
        celda_inferior3 = hoja.cell(row=celda_inicio+3, column=1, value="T. Canales - T. Argentina")
        celda_inferior4 = hoja.cell(row=celda_inicio+4, column=1, value="* Baja muestra")


    celda_inferior1.font = fuente_negrita
    celda_inferior2.font = fuente_negrita
    celda_inferior3.font = fuente_negrita
    celda_inferior4.font = fuente_roja 

    ##FUENTE A LAS CELDAS
    num_filas = hoja.max_row
    num_columnas = hoja.max_column

    ##validar 
    lista_de_palabras = ["% Volumen", "Penetración (%)","Compra media (kg)","Compra por acto (kg)","Frecuencia (veces)"]
    numeros_filas_activas = []
    
    # Iterar sobre todas las filas de la hoja
    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=1):
        celda = fila[0]  # Obtener la celda de la primera columna en la fila
        palabra = celda.value  # Obtener el valor de la celda
        # Verificar si la celda contiene una palabra que te interesa
        if palabra in lista_de_palabras:
            numero_fila = celda.row  # Obtener el número de la fila
            numeros_filas_activas.append(numero_fila)  # Agregar el número de fila a la lista

    fila_inicial = numeros_filas_activas[0]  # Por ejemplo, si el texto "Dif vs PY" está en la fila 2
    ##validar

    solo_negrita,solo_azul_sin_negrita,solo_negrita_azul,negrita_fondo_gris = lista_de_sombreoc(nombre_hoja)  # Lista de palabras a buscar
    # Listas para almacenar ubicaciones según el número de asteriscos

    primer_filtro = []  # Lista para una estrella
    segundo_filtro = []  # Lista para dos estrellas
    tercer_filtro = []  # Lista para tres estrellas
    cuarto_filtro = []  # Lista para cuatro estrellas

    ##AQUI ARREGLO LOS COLORES 24/04

    primer_filtro = []  # Lista para una estrella - solo_negrita
    segundo_filtro = []  # Lista para dos estrellas - solo_azul_sin_negrita
    tercer_filtro = []  # Lista para tres estrellas - solo_negrita_azul
    cuarto_filtro = []  # Lista para cuatro estrellas - negrita_fondo_gris
    relleno_azul = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    for fila in hoja.iter_rows(min_row=fila_inicial, max_row=num_filas, min_col=1, max_col=1):
        celda = fila[0]  # Obtener la celda de la primera columna en la fila
        palabra = celda.value  # Obtener el valor de la celda
        if palabra in solo_negrita:
            numero_fila = celda.row  # Obtener el número de la fila
            primer_filtro.append(numero_fila)  # Agregar el número de fila a la lista - solo_negrita
        elif palabra in solo_azul_sin_negrita:
            numero_fila = celda.row  # Obtener el número de la fila
            segundo_filtro.append(numero_fila)  # Agregar el número de fila a la lista - solo_azul_sin_negrita
        elif palabra in solo_negrita_azul:
            numero_fila = celda.row  # Obtener el número de la fila
            tercer_filtro.append(numero_fila)  # Agregar el número de fila a la lista - solo_negrita_azul
        elif palabra in negrita_fondo_gris:
            numero_fila = celda.row  # Obtener el número de la fila
            cuarto_filtro.append(numero_fila)  # Agregar el número de fila a la lista - negrita_fondo_gris

    
    # Formato de relleno para aplicar
    relleno_azul = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    relleno_gris = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    # Iterar sobre las columnas activas de las filas guardadas en filas_palabras y aplicar el formato de relleno
    
    if primer_filtro:
        for fila_ubicacion in primer_filtro:
            for columna in hoja.iter_cols(min_row=fila_ubicacion, max_row=fila_ubicacion):
                for celda in columna:
                    if celda.value != None and celda.value != "":  # Verificar si la celda está activa (no es una celda combinada)
                        #celda.fill = relleno_azul 
                        celda.font = fuente_negrita


    if segundo_filtro:
        for fila_ubicacion in segundo_filtro:
            for columna in hoja.iter_cols(min_row=fila_ubicacion, max_row=fila_ubicacion):
                for celda in columna:
                    if celda.value != None and celda.value != "":  # Verificar si la celda está activa (no es una celda combinada)
                        #celda.fill = relleno_gris
                        #celda.font = Font(color='000000', size=10, name='Arial', bold=True)
                        celda.fill = relleno_azul 

    if tercer_filtro:    
        for fila_ubicacion in tercer_filtro:
            for columna in hoja.iter_cols(min_row=fila_ubicacion, max_row=fila_ubicacion):
                for celda in columna:
                    if celda.value != None and celda.value != "":  # Verificar si la celda está activa (no es una celda combinada)
                        celda.fill = relleno_azul
                        celda.font = fuente_negrita
    if cuarto_filtro:
        for fila_ubicacion in cuarto_filtro:
            for columna in hoja.iter_cols(min_row=fila_ubicacion, max_row=fila_ubicacion):
                for celda in columna:
                    if celda.value != None and celda.value != "":  # Verificar si la celda está activa (no es una celda combinada)
                        celda.fill = relleno_gris
                        celda.font = Font(color='000000', size=8, name='Arial', bold=True)


##FOrmatos
def aplicar_formato(hoja,nombre_hoja):
    lista_de_palabras = ["Penetración (%)","Compra media (kg)","Compra por acto (kg)","Frecuencia (veces)"]
    numeros_filas_activas = []
    
    # Iterar sobre todas las filas de la hoja
    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=1):
        celda = fila[0]  # Obtener la celda de la primera columna en la fila
        palabra = celda.value  # Obtener el valor de la celda
        # Verificar si la celda contiene una palabra que te interesa
        if palabra in lista_de_palabras:
            numero_fila = celda.row  # Obtener el número de la fila
            numeros_filas_activas.append(numero_fila)  # Agregar el número de fila a la lista

    fuente_negrita_blanco = Font(color='FFFFFFFF', name='Arial', size=8, bold=True)
    relleno = PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid')  # Color rojo en este ejemplo
    
    #23/04
    fuente_negrita_generico = Font(color='000000', name='Arial', size=8)
    num_filas = hoja.max_row
    num_columnas = hoja.max_column
    for fila in hoja.iter_rows(min_row=1, max_row=num_filas, min_col=1, max_col=num_columnas):
        #if numero_fila not in numeros_filas_activas:
        for celda in fila:
            if celda.value is not None and isinstance(celda.value, (str, int, float)):
                celda.font = fuente_negrita_generico

    # Aplicar formato a las celdas activas
    for numero_fila in numeros_filas_activas:
        for columna in hoja.iter_cols(min_row=numero_fila, max_row=numero_fila, min_col=1, max_col=hoja.max_column):
            for celda in columna:
                if celda.value is not None and isinstance(celda.value, str) and celda.value.strip() != "": # Verificar que la celda no esté vacía
                    celda.fill = relleno  
                    celda.font = fuente_negrita_blanco
    
    # Obtener el rango de celdas de las filas activas
    rango_celdas = f"A{numeros_filas_activas[0]}:AZ{numeros_filas_activas[-1]}"
    celdas_filas_activas = hoja[rango_celdas]

    # Aplicar los formatos a todas las celdas del rango
    for fila in celdas_filas_activas:
        for celda in fila:
            celda.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Estilo de borde blanco
    borde_blanco = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
    )

    # Borde blanco a todo 
    for fila in hoja.iter_rows(min_row=1, max_row=300, min_col=1, max_col=300):
        for celda in fila:
            celda.border = borde_blanco

    # Crear el formato para el borde negro
    borde_negro = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))

    num_filas = hoja.max_row
    num_columnas = hoja.max_column

    # Aplicar bordes y formatos numéricos
    for numero_fila, fila in enumerate(hoja.iter_rows(min_row=1, max_row=num_filas, min_col=1, max_col=num_columnas), start=1):
        #if numero_fila not in numeros_filas_activas:
            for celda in fila:
                valor = celda.value
                if  isinstance(valor, str):
                    if valor is not None and valor.strip() == "":
                        celda.border = borde_blanco
                    else:
                        celda.border = borde_negro
                elif isinstance(valor, (int, float)):
                    celda.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    celda.border = borde_negro

    hoja.column_dimensions["A"].width = 26.60
    for numero_fila, fila in enumerate(hoja.iter_rows(min_row=1, max_row=num_filas, min_col=1, max_col=num_columnas), start=1):
        if numero_fila not in numeros_filas_activas:
            for celda in fila:
                valor = celda.value
                if isinstance(valor, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0.0' ##29/04
                else: 
                    celda.alignment = Alignment(horizontal='left', vertical='center')

    # Inicializar la lista para almacenar las columnas que contienen "Dif vs PY"
    columnas_dif_vs_py = []

    # Definir la fila desde la cual comenzar la búsqueda
    fila_inicial = numeros_filas_activas[0]  # Por ejemplo, si el texto "Dif vs PY" está en la fila 2

    # Iterar sobre las columnas de la hoja
    for columna in hoja.iter_cols(min_row=fila_inicial, max_row=fila_inicial):
        texto_columna = columna[0].value  # Obtener el texto de la primera celda de la columna
        if texto_columna is not None and "Dif vs PY" in texto_columna.strip():  # Verificar si la columna contiene "Dif vs PY"
            columnas_dif_vs_py.append(columna[0].column)  # Guardar la ubicación de la columna

    ##cambio &&
    #penetracion
    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[0],max_row=numeros_filas_activas[1] -1 , min_col=1, max_col=num_columnas):
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0.0'  # Formato de porcentaje con dos decimales - penetracion es un decimal 


    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[1], max_row=numeros_filas_activas[-1] -1, min_col=col_num, max_col=col_num):
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0%'  # Formato de porcentaje con dos decimales - compra media y compra x accto
    
    columnas_no_en_col_num = [col for col in range(1, num_columnas + 1) if col not in columnas_dif_vs_py]
    for col_num in columnas_no_en_col_num:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[-1], min_col=col_num, max_col=col_num): #formato geo 2decimales
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0.0'  # Formato de porcentaje con  decimales - frecuencia

    
    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[-1], min_col=col_num, max_col=col_num):
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0%'  # Formato de porcentaje con dos decimales - frecuencia

    ##AJUSTAR COLUMNAS NO DIF YS 25/04

    # Obtener el índice de la última columna
    ultima_columna = hoja.max_column

    # Iterar sobre las columnas guardadas y ajustar el tamaño de las columnas a la derecha
    for columna_idx in columnas_dif_vs_py:
        if columna_idx < ultima_columna:
            hoja.column_dimensions[get_column_letter(columna_idx + 1)].width = 3  # Ajustar el tamaño de la columna siguiente

    #27/04
    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[0], max_row=numeros_filas_activas[1] - 1, min_col=1, max_col=num_columnas):
            for index, celda in enumerate(fila):
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    if celda.value < 0 and fila[index - 1].value < 3:
                        fila[index - 1].font = Font(color='FF0000', name='Arial', size=8)
    
    #Penetracion menor a 3 puntos (baja muestra)  25/07/2024
    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[0], max_row=numeros_filas_activas[1] - 1, min_col=col_num-1, max_col=col_num-1):
            for index, celda in enumerate(fila):
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    if fila[index - 1].value < 3:
                        fila[index - 1].font = Font(color='FF0000', name='Arial', size=8)
    
def aplicar_tamaño_regiones_canales(hoja,nombre_hoja):
    lista_de_palabras = ["% Volumen", "Penetración (%)","Compra media (kg)","Compra por acto (kg)","Frecuencia (veces)"]
    numeros_filas_activas = []
    # Iterar sobre todas las filas de la hoja
    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=1):
        celda = fila[0]  # Obtener la celda de la primera columna en la fila
        palabra = celda.value  # Obtener el valor de la celda
        # Verificar si la celda contiene una palabra que te interesa
        if palabra in lista_de_palabras:
            numero_fila = celda.row  # Obtener el número de la fila
            numeros_filas_activas.append(numero_fila)  # Agregar el número de fila a la lista

    # Inicializar la lista para almacenar las columnas que contienen "Dif vs PY"
    columnas_dif_vs_py = []
    # Definir la fila desde la cual comenzar la búsqueda
    fila_inicial = numeros_filas_activas[0]  # Por ejemplo, si el texto "Dif vs PY" está en la fila 2

    # Iterar sobre las columnas de la hoja
    for columna in hoja.iter_cols(min_row=fila_inicial, max_row=fila_inicial):
        texto_columna = columna[0].value  # Obtener el texto de la primera celda de la columna
        if texto_columna is not None and "Dif vs PY" in texto_columna.strip():  # Verificar si la columna contiene "Dif vs PY"
            columnas_dif_vs_py.append(columna[0].column)  # Guardar la ubicación de la columna

    celda_tamaño_real=[]
        # Iterar sobre las columnas de la hoja
    for columna in hoja.iter_cols(min_row=fila_inicial, max_row=fila_inicial):
        texto_columna = columna[0].value  # Obtener el texto de la primera celda de la columna
        if texto_columna is not None and  texto_columna.strip() in lista_de_palabras:  # Verificar si la columna contiene "Dif vs PY"
            celda_tamaño_real.append(columna[0].column)  # Guardar la ubicación de la 

    for col in celda_tamaño_real:
        col_entero = int(col)
        columna = get_column_letter(col_entero)
        hoja.column_dimensions[columna].width = 26.60
   
def aplicar_formato_geo(hoja,nombre_hoja):
    lista_de_palabras = ["% Volumen", "Penetración (%)","Compra media (kg)","Compra por acto (kg)","Frecuencia (veces)"]
    numeros_filas_activas = []
    
    # Iterar sobre todas las filas de la hoja
    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=1):
        celda = fila[0]  # Obtener la celda de la primera columna en la fila
        palabra = celda.value  # Obtener el valor de la celda
        # Verificar si la celda contiene una palabra que te interesa
        if palabra in lista_de_palabras:
            numero_fila = celda.row  # Obtener el número de la fila
            numeros_filas_activas.append(numero_fila)  # Agregar el número de fila a la lista

    fuente_negrita_blanco = Font(color='FFFFFFFF', name='Arial', size=8, bold=True)
    relleno = PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid')  # Color rojo en este ejemplo
    

    #23/04
    fuente_negrita_generico = Font(color='000000', name='Arial', size=8)
    num_filas = hoja.max_row
    num_columnas = hoja.max_column
    for fila in hoja.iter_rows(min_row=1, max_row=num_filas, min_col=1, max_col=num_columnas):
        #if numero_fila not in numeros_filas_activas:
        for celda in fila:
            if celda.value is not None and isinstance(celda.value, (str, int, float)):
                celda.font = fuente_negrita_generico

    # Aplicar formato a las celdas activas
    for numero_fila in numeros_filas_activas:
        for columna in hoja.iter_cols(min_row=numero_fila, max_row=numero_fila, min_col=1, max_col=hoja.max_column):
            for celda in columna:
                if celda.value is not None and isinstance(celda.value, str) and celda.value.strip() != "": # Verificar que la celda no esté vacía
                    celda.fill = relleno  
                    celda.font = fuente_negrita_blanco
    
    # Obtener el rango de celdas de las filas activas
    rango_celdas = f"A{numeros_filas_activas[0]}:AZ{numeros_filas_activas[-1]}"
    celdas_filas_activas = hoja[rango_celdas]

    # Aplicar los formatos a todas las celdas del rango
    for fila in celdas_filas_activas:
        for celda in fila:
            celda.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Estilo de borde blanco
    borde_blanco = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
    )

    # Borde blanco a todo 
    for fila in hoja.iter_rows(min_row=1, max_row=300, min_col=1, max_col=300):
        for celda in fila:
            celda.border = borde_blanco

    # Crear el formato para el borde negro
    borde_negro = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))

    num_filas = hoja.max_row
    num_columnas = hoja.max_column

    # Aplicar bordes y formatos numéricos
    for numero_fila, fila in enumerate(hoja.iter_rows(min_row=1, max_row=num_filas, min_col=1, max_col=num_columnas), start=1):
        #if numero_fila not in numeros_filas_activas:
            for celda in fila:
                valor = celda.value
                if  isinstance(valor, str):
                    if valor is not None and valor.strip() == "":
                        celda.border = borde_blanco
                    else:
                        celda.border = borde_negro
                elif isinstance(valor, (int, float)):
                    celda.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    celda.border = borde_negro

    hoja.column_dimensions["A"].width = 26.60
    for numero_fila, fila in enumerate(hoja.iter_rows(min_row=1, max_row=num_filas, min_col=1, max_col=num_columnas), start=1):
        if numero_fila not in numeros_filas_activas:
            for celda in fila:
                valor = celda.value
                if isinstance(valor, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0.0' #29/04
                else: 
                    celda.alignment = Alignment(horizontal='left', vertical='center')

    # Inicializar la lista para almacenar las columnas que contienen "Dif vs PY"
    columnas_dif_vs_py = []

    # Definir la fila desde la cual comenzar la búsqueda
    fila_inicial = numeros_filas_activas[0]  # Por ejemplo, si el texto "Dif vs PY" está en la fila 2

    # Iterar sobre las columnas de la hoja
    for columna in hoja.iter_cols(min_row=fila_inicial, max_row=fila_inicial):
        texto_columna = columna[0].value  # Obtener el texto de la primera celda de la columna
        if texto_columna is not None and "Dif vs PY" in texto_columna.strip():  # Verificar si la columna contiene "Dif vs PY"
            columnas_dif_vs_py.append(columna[0].column)  # Guardar la ubicación de la columna

    ## cambio && geo

    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[0], max_row=numeros_filas_activas[2] -1, min_col=1, max_col=num_columnas): #formato geo 2decimales
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0.0'  # Formato de porcentaje con dos decimales - volumen - penetracioj

    #27/04
    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[1], max_row=numeros_filas_activas[2] - 1, min_col=1, max_col=num_columnas):
            for index, celda in enumerate(fila):
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    if celda.value < 0 and fila[index - 1].value < 3:
                        fila[index - 1].font = Font(color='FF0000', name='Arial', size=8)

    #PENETRACION MENOR A 3 PUNTOS <baja muestra> 25/07/2024
    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[1], max_row=numeros_filas_activas[2] - 1, min_col=col_num-1, max_col=col_num-1):
            for index, celda in enumerate(fila):
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    if fila[index - 1].value < 3:
                        fila[index - 1].font = Font(color='FF0000', name='Arial', size=8)


    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[2],max_row=numeros_filas_activas[-1] -1, min_col=col_num, max_col=col_num): #formato geo 2decimales
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0%'  # Formato de porcentaje con dos decimales - compra media . compra x acto

    columnas_no_en_col_num = [col for col in range(1, num_columnas + 1) if col not in columnas_dif_vs_py]
    for col_num in columnas_no_en_col_num:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[-1], min_col=col_num, max_col=col_num): #formato geo 2decimales
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0.0'  # Formato de porcentaje con dos decimales - frecuencia

    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[-1], min_col=col_num, max_col=col_num): #formato geo 2decimales
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0%'  # Formato de porcentaje con dos decimales - frecuencia
    

    celda_tamaño_real=[]
        # Iterar sobre las columnas de la hoja
    for columna in hoja.iter_cols(min_row=fila_inicial, max_row=fila_inicial):
        texto_columna = columna[0].value  # Obtener el texto de la primera celda de la columna
        if texto_columna is not None and "% Volumen" in texto_columna.strip():  # Verificar si la columna contiene "Dif vs PY"
            celda_tamaño_real.append(columna[0].column)  # Guardar la ubicación de la 

    for col in celda_tamaño_real:
        col_entero = int(col)
        columna = get_column_letter(col_entero)
        hoja.column_dimensions[columna].width = 26.60

    ##AJUSTAR COLUMNAS NO DIF YS 25/04

    # Obtener el índice de la última columna
    ultima_columna = hoja.max_column

    # Iterar sobre las columnas guardadas y ajustar el tamaño de las columnas a la derecha
    for columna_idx in columnas_dif_vs_py:
        if columna_idx < ultima_columna:
            hoja.column_dimensions[get_column_letter(columna_idx + 1)].width = 3  # Ajustar el tamaño de la columna siguiente
              
def aplicar_formato_region_segmento(hoja,nombre_hoja):
    lista_de_palabras = ["% Volumen", "Penetración (%)","Compra media (kg)","Compra por acto (kg)","Frecuencia (veces)"]
    numeros_filas_activas = []
    
    # Iterar sobre todas las filas de la hoja
    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=1):
        celda = fila[0]  # Obtener la celda de la primera columna en la fila
        palabra = celda.value  # Obtener el valor de la celda
        # Verificar si la celda contiene una palabra que te interesa
        if palabra in lista_de_palabras:
            numero_fila = celda.row  # Obtener el número de la fila
            numeros_filas_activas.append(numero_fila)  # Agregar el número de fila a la lista

    fuente_negrita_blanco = Font(color='FFFFFFFF', name='Arial', size=8, bold=True)
    relleno = PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid')  # Color rojo en este ejemplo
    
    #23/04
    fuente_negrita_generico = Font(color='000000', name='Arial', size=8)
    num_filas = hoja.max_row
    num_columnas = hoja.max_column
    for fila in hoja.iter_rows(min_row=1, max_row=num_filas, min_col=1, max_col=num_columnas):
        #if numero_fila not in numeros_filas_activas:
        for celda in fila:
            if celda.value is not None and isinstance(celda.value, (str, int, float)):
                celda.font = fuente_negrita_generico
    # Aplicar formato a las celdas activas
    for numero_fila in numeros_filas_activas:
        for columna in hoja.iter_cols(min_row=numero_fila, max_row=numero_fila, min_col=1, max_col=hoja.max_column):
            for celda in columna:
                if celda.value is not None and isinstance(celda.value, str) and celda.value.strip() != "": # Verificar que la celda no esté vacía
                    celda.fill = relleno  
                    celda.font = fuente_negrita_blanco
    
    # Obtener el rango de celdas de las filas activas
    rango_celdas = f"A{numeros_filas_activas[0]}:AZ{numeros_filas_activas[-1]}"
    celdas_filas_activas = hoja[rango_celdas]

    # Aplicar los formatos a todas las celdas del rango
    for fila in celdas_filas_activas:
        for celda in fila:
            celda.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Estilo de borde blanco
    borde_blanco = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
    )

    # Borde blanco a todo 
    for fila in hoja.iter_rows(min_row=1, max_row=300, min_col=1, max_col=300):
        for celda in fila:
            celda.border = borde_blanco

    # Crear el formato para el borde negro
    borde_negro = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))

    num_filas = hoja.max_row
    num_columnas = hoja.max_column

    # Aplicar bordes y formatos numéricos
    for numero_fila, fila in enumerate(hoja.iter_rows(min_row=1, max_row=num_filas, min_col=1, max_col=num_columnas), start=1):
        #if numero_fila not in numeros_filas_activas:
            for celda in fila:
                valor = celda.value
                if  isinstance(valor, str):
                    if valor is not None and valor.strip() == "":
                        celda.border = borde_blanco
                    else:
                        celda.border = borde_negro
                elif isinstance(valor, (int, float)):
                    celda.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    celda.border = borde_negro

    hoja.column_dimensions["A"].width = 26.60
    for numero_fila, fila in enumerate(hoja.iter_rows(min_row=1, max_row=num_filas, min_col=1, max_col=num_columnas), start=1):
        if numero_fila not in numeros_filas_activas:
            for celda in fila:
                valor = celda.value
                if isinstance(valor, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0.0'#29/04
                else: 
                    celda.alignment = Alignment(horizontal='left', vertical='center')

    # Inicializar la lista para almacenar las columnas que contienen "Dif vs PY"
    columnas_dif_vs_py = []

    # Definir la fila desde la cual comenzar la búsqueda
    fila_inicial = numeros_filas_activas[0]  # Por ejemplo, si el texto "Dif vs PY" está en la fila 2

    # Iterar sobre las columnas de la hoja
    for columna in hoja.iter_cols(min_row=fila_inicial, max_row=fila_inicial):
        texto_columna = columna[0].value  # Obtener el texto de la primera celda de la columna
        if texto_columna is not None and "Dif vs PY" in texto_columna.strip():  # Verificar si la columna contiene "Dif vs PY"
            columnas_dif_vs_py.append(columna[0].column)  # Guardar la ubicación de la columna


    ##cambio &&
    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[0],max_row=numeros_filas_activas[1] -1 , min_col=1, max_col=num_columnas):
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0.0'  # Formato de porcentaje con dos decimales - penetracion es un decimal 

    #27/04
    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[0], max_row=numeros_filas_activas[1] - 1, min_col=1, max_col=num_columnas):
            for index, celda in enumerate(fila):
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    if celda.value < 0 and fila[index - 1].value < 3:
                        fila[index - 1].font = Font(color='FF0000', name='Arial', size=8)
    

    #Penetracion menor a 3 puntos (baja muestra)  25/07/2024
    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[0], max_row=numeros_filas_activas[1] - 1, min_col=col_num-1, max_col=col_num-1):
            for index, celda in enumerate(fila):
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    if fila[index - 1].value < 3:
                        fila[index - 1].font = Font(color='FF0000', name='Arial', size=8)

    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[1],max_row=numeros_filas_activas[-1] -1 , min_col=col_num, max_col=col_num): #formato geo 2decimales
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0%'  # Formato de porcentaje con dos decimales - compra - compraxacto

    columnas_no_en_col_num = [col for col in range(1, num_columnas + 1) if col not in columnas_dif_vs_py]
    for col_num in columnas_no_en_col_num:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[-1], min_col=col_num, max_col=col_num): #formato geo 2decimales
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0.0'  # Formato de porcentaje con dos decimales - frecuencia

    for col_num in columnas_dif_vs_py:
        for fila in hoja.iter_rows(min_row=numeros_filas_activas[-1], min_col=col_num, max_col=col_num): #formato geo 2decimales
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0%'  # Formato de porcentaje con dos decimales - frecuencia

    celda_tamaño_real=[]
        # Iterar sobre las columnas de la hoja
    for columna in hoja.iter_cols(min_row=fila_inicial, max_row=fila_inicial):
        texto_columna = columna[0].value  # Obtener el texto de la primera celda de la columna
        if texto_columna is not None and "% Volumen" in texto_columna.strip():  # Verificar si la columna contiene "Dif vs PY"
            celda_tamaño_real.append(columna[0].column)  # Guardar la ubicación de la 

    for col in celda_tamaño_real:
        col_entero = int(col)
        columna = get_column_letter(col_entero)
        hoja.column_dimensions[columna].width = 26.60
    {}
    ##AJUSTAR COLUMNAS NO DIF YS 25/04

    # Obtener el índice de la última columna
    ultima_columna = hoja.max_column

    # Iterar sobre las columnas guardadas y ajustar el tamaño de las columnas a la derecha
    for columna_idx in columnas_dif_vs_py:
        if columna_idx < ultima_columna:
            hoja.column_dimensions[get_column_letter(columna_idx + 1)].width = 3  # Ajustar el tamaño de la columna siguiente
            

def indices_marca_yogures_postres_queso(nombre_hoja):

    Indices = {
        "L_REGION = Total Argentina\Total L_REGION":"Principal",		
        "WPOBRYG = T. Yogurisimo\Total WPOBRYG":"Resto Yogurisimo",			
        "WPOBRYG = T. Ser\Total WPOBRYG":"Resto Ser",	
        "WPOBRYG = LS Clasico (Danone)\Total WPOBRYG":"LS Clasico (Danone)"	}																						
    
    Indices2 = {
        "COMBQE3 = T. Quesos Blancos + Fundidos\Total COMBQE3":"Principal",		
        "COMBQE3 = T. Quesos Blancos + Fundidos\Total COMBQE3 -Kantar":"Cremón"}																						

    subcadenas = nombre_hoja.split("_") 
    if subcadenas[1]=="yog": 
        return Indices
    else: 
        return Indices2

#Lectura 
def obtener_nombres_hojas(direccion_base1):
    nombres = []
    # Carga el archivo Excel
    wb = openpyxl.load_workbook(direccion_base1)

    # Obtiene los nombres de las hojas
    nombres_hojas = wb.sheetnames
    # Itera a través de los nombres de las hojas
    for nombre in nombres_hojas:
        nombres.append(nombre)
    return nombres 


def summary_funcion(direccion_base2 ):
        
    def obtener_nombres_hojas_summary(direccion_base1):
        nombres = []
        # Carga el archivo Excel
        wb = openpyxl.load_workbook(direccion_base1)

        # Obtiene los nombres de las hojas
        nombres_hojas = wb.sheetnames
        # Itera a través de los nombres de las hojas
        for nombre in nombres_hojas:
            if "Marcas" in nombre:
                nombres.append(nombre)
        return nombres 

    nombres_hojas_summary = obtener_nombres_hojas_summary(direccion_base2)

    ## OBTENER SOBRE TITULO 
    def obtener_titulo(nombre):
        if "yog" in nombre:
            return  "YOGURES"
        elif "QE" in nombre: 
            return "QUESOS BLANCOS + FUNDIDOS"
        elif "PO" in nombre: 
            return "POSTRES"

    lista_eliminacion = [
        "YOGURES",
        "  T. Yogur + Transito + Salud  ",
        "    T. Danone  ",
        "     T. Yogurisimo  ",
        "     T. Ser  ",
        "     LS Clasico (Danone)  ",
        "     Gran Compra  ",
        "     Danonino  ",
        "    Actimel  ",
        "QUESOS BLANCOS + FUNDIDOS",
        "  T.Quesos Blancos + Fundidos  ",
        "   T. Danone  ",
        "    Casancrem  ",
        "    LS Clásico  ",
        "POSTRES",
        "  T. Pos + Fla + Gel + Azl  ",
        "   T. Danone  ",
        "    T. Danette  ",
        "    T. Ser  ",
        "    T. Serenito  "
    ]

    lista_sobreado_summary =[ 
        "  T. Yogur + Transito + Salud  ",
        "    T. Danone  ",
        "  T.Quesos Blancos + Fundidos  ",
        "   T. Danone  ", 
        "  T. Pos + Fla + Gel + Azl  ",
        "   T. Danone  "
        ]




    # Cargar el libro de Excel existente
    wb = load_workbook(filename=direccion_base2)


    def conseguir_summary_uno(nombre):
        hoja = wb[nombre]
        lista_de_palabras = ["Penetración (%)","Compra media (kg)","Compra por acto (kg)","Frecuencia (veces)"]
        numeros_filas_activas = []
        numeros_filas_medidas = {}
            
        # Iterar sobre todas las filas de la hoja
        for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=1):
            celda = fila[0]  # Obtener la celda de la primera columna en la fila
            palabra = celda.value  # Obtener el valor de la celda
            # Verificar si la celda contiene una palabra que te interesa
            if palabra in lista_de_palabras:
                numero_fila = celda.row  # Obtener el número de la fila
                numeros_filas_activas.append(numero_fila)  # Agregar el número de fila a la lista
                numeros_filas_medidas[palabra]={numero_fila}


        # Ejemplo de cómo convertir un conjunto a un diccionario con las mismas claves y valores vacíos


        # Crear un diccionario para almacenar los DataFrames separados
        dataframes_separados = {}

        # Iterar sobre los números de fila para crear los DataFrames separados
        for i in range(len(numeros_filas_activas)):
            # Obtener el número de fila de inicio del indicador actual
            fila_inicio = numeros_filas_activas[i]
            
            # Calcular el número de fila de inicio del siguiente indicador
            if i + 1 < len(numeros_filas_activas):
                fila_siguiente_inicio = numeros_filas_activas[i + 1]
            else:
                # Si es el último indicador, usar la última fila del DataFrame
                fila_siguiente_inicio = hoja.max_row + 1
            
            # Crear un DataFrame con los datos desde fila_inicio hasta fila_siguiente_inicio - 1
            datos_indicador = [[celda.value for celda in fila] for fila in hoja.iter_rows(min_row=fila_inicio, max_row=fila_siguiente_inicio - 1, min_col=1, max_col=hoja.max_column)]
            df_indicador = DataFrame(datos_indicador)
            
            # Eliminar las columnas con valores None
            #df = df_indicador.dropna(axis=1, how='all')
            #df = df.dropna(axis=0, how='all')
            #insetar aquí
            df = df_indicador.dropna(axis=0, how='all')
            titulo = obtener_titulo(nombre)
            df.columns = df.iloc[0]  # Convertir la primera fila en nombres de columnas
            df = df.iloc[1:]  # Eliminar la primera fila (ahora es el encabezado)
            #df = df.reset_index(drop=True)  # Reiniciar los índices
            df.reset_index(drop=True, inplace=True) ## eliminar aquí
            # Obtener el nombre de la primera columna del DataFrame
            primer_nombre_columna = df.columns[0]
            df.columns = [titulo] + df.columns[1:].tolist()  # Cambiar el nombre de la primera columna
            # Agregar el DataFrame al diccionario de DataFrames separados
            #dataframes_separados[f"primer_nombre_columna"] = df
            df= df[df[titulo].isin(lista_eliminacion)]
            dataframes_separados[primer_nombre_columna] = df
            # Eliminar el índice del DataFrame



        return dataframes_separados


    lista_directorio_dataframe = {}
    for nombre in nombres_hojas_summary: 
        data = conseguir_summary_uno(nombre)
        lista_directorio_dataframe[nombre] = data

    data1 = conseguir_summary_uno("Marcas_yog")
    data2 = conseguir_summary_uno("Marcas_QE")
    data3 = conseguir_summary_uno("Marcas_PO")


    lista_datos = [data1, data2, data3]
    lista_de_medidas = ["Penetración (%)","Compra media (kg)","Compra por acto (kg)","Frecuencia (veces)"]

    nombre_hoja = "Summary Danone"
    # Nombre del archivo Excel que quieres crear
    # Crear una nueva hoja en el libro de Excel/ 'nombre_de_la_nueva_hoja'
    ws = wb.create_sheet(nombre_hoja) 

    last_row = 6
    for medida in lista_de_medidas: 

        for dataframes in lista_datos:     

            rows = dataframe_to_rows(dataframes[medida], index=False, header=True)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=last_row + r_idx, column=c_idx, value=value)

                # Insertar una celda vacía como separación en la columna A
            ws.cell(row=last_row + r_idx, column=1)

            # Actualizar la última fila ocupada en la hoja de trabajo
            last_row = ws.max_row
        last_row = ws.max_row+3

    nombre_del_mes = fecha_formateada
    celda_inicio=1
    fuente_negrita = Font(color='000000', name='Arial', size=8, bold=True)
    fuente_roja =  Font(color='FF0000', name='Arial', size=8)

    # Estilo de borde blanco
    borde_blanco = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
    )

        # Crear el formato para el borde negro
    borde_negro = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))
    
    # Borde blanco a todo 
    for fila in ws.iter_rows(min_row=1, max_row=300, min_col=1, max_col=300):
        for celda in fila:
            celda.border = borde_blanco


    lista_filas_resaltadas = ['YOGURES','QUESOS BLANCOS + FUNDIDOS','POSTRES']
    numeros_filas_a_resaltar = []
    # Iterar sobre todas las filas de la hoja
    for fila in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        celda = fila[0]  # Obtener la celda de la primera columna en la fila
        palabra = celda.value  # Obtener el valor de la celda
        # Verificar si la celda contiene una palabra que te interesa
        if palabra in lista_filas_resaltadas:
            numero_fila = celda.row  # Obtener el número de la fila
            numeros_filas_a_resaltar.append(numero_fila)  # Agregar el número de fila a la lista

    num_filas = ws.max_row
    num_columnas = ws.max_column


    fuente_negrita_blanco = Font(color='FFFFFFFF', name='Arial', size=8, bold=True)
    fuente_negrita_generico = Font(color='000000', name='Arial', size=8)

    relleno = PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid')  # Color rojo en este ejemplo
    
    ####APLICA
    # Aplicar fORMATO A INDICES YOG-QUESOS-POSTRES
    for num_filas in numeros_filas_a_resaltar:
        for fila in ws.iter_rows(min_row=num_filas, max_row=num_filas,max_col=num_columnas):
            #if numero_fila not in numeros_filas_activas:
                for celda in fila:
                    if celda.value is not None and isinstance(celda.value, str) and celda.value.strip() != "": # Verificar que la celda no esté vacía
                        celda.fill = relleno
                        celda.font = fuente_negrita_blanco
                        celda.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ##TITULOS DE LOS INDICADORES 


    ws.column_dimensions["A"].width = 26.60
    
    # Supongamos también que tienes el número total de filas en tu hoja de Excel
    num_filas_total = num_filas
    # Filas que no están en la lista de filas a resaltar
    filas_no_en_fil_num = [fila for fila in range(1, num_filas_total + 10) if fila not in numeros_filas_a_resaltar]

    for no_filas in filas_no_en_fil_num: 
        for fila in ws.iter_rows(min_row=no_filas, max_row=no_filas,max_col=num_columnas):
        #if numero_fila not in numeros_filas_activas:
            for celda in fila:
                if celda.value is not None and isinstance(celda.value, (str, int, float)):
                    celda.font = fuente_negrita_generico
  
            

    # Inicializar la lista para almacenar las columnas que contienen "Dif vs PY"
    columnas_dif_vs_py = []

    # Definir la fila desde la cual comenzar la búsqueda
    fila_inicial = numeros_filas_a_resaltar[0]  # Por ejemplo, si el texto "Dif vs PY" está en la fila 2
    
    # Iterar sobre las columnas de la hoja
    for columna in ws.iter_cols(min_row=fila_inicial, max_row=fila_inicial):
        texto_columna = columna[0].value  # Obtener el texto de la primera celda de la columna
        if texto_columna is not None and "Dif vs PY" in texto_columna.strip():  # Verificar si la columna contiene "Dif vs PY"
            columnas_dif_vs_py.append(columna[0].column)  # Guardar la ubicación de la columna

    #Columnas que no tienen dif 
    columnas_no_en_col_num = [col for col in range(1, num_columnas + 1) if col not in columnas_dif_vs_py]
    

    #Formato de decimales a las columnas con dif
    for col_num in columnas_dif_vs_py:
        for fila in ws.iter_rows(min_row=27, min_col=col_num, max_col=col_num):
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0%'  # Formato de porcentaje con dos decimales
                    celda.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    #Formato de decimales a las columnas que con dif pero solo penetracion 
    for col_num in columnas_dif_vs_py:
        for fila in ws.iter_rows(min_row=4, max_row=26, min_col=col_num, max_col=col_num):
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0.0'  # Formato de porcentaje con dos decimales
                    celda.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center') 

    #27/04 penetracion

    #27/04 - summary
    for col_num in columnas_dif_vs_py:
        for fila in ws.iter_rows(min_row=4, max_row=26, min_col=1, max_col=num_columnas):
            for index, celda in enumerate(fila):
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    if celda.value < 0 and fila[index - 1].value < 3:
                        fila[index - 1].font = Font(color='FF0000', name='Arial', size=8)
    

    #Formato para las que no on dif en todas las linesas
        #Formato de decimales a las columnas con dif
    for col_num in columnas_no_en_col_num:
        for fila in ws.iter_rows(min_row=5, min_col=col_num, max_col=col_num):
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.value = round(celda.value, 2)
                    celda.number_format = '0.0'  # Formato de porcentaje con dos decimales
                    celda.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        


    fuente_negrita = Font(color='000000', name='Arial', size=8, bold=True)

    ##COMIENZO A PEGAR EL FORMATO POR NOMBRES
    
    lista_azul_summary=[]
    relleno_azul = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    for fila in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        celda = fila[0]  # Obtener la celda de la primera columna en la fila
        palabra = celda.value  # Obtener el valor de la celda
        if palabra in lista_sobreado_summary:
            #celda.fill = relleno_azul
            #celda.font = fuente_negrita
            numero_fila = celda.row  # Obtener el número de la fila
            lista_azul_summary.append(numero_fila)  # Agregar el número de fila a la lista

    ##AJUSTAR COLUMNAS NO DIF YS 25/04

    # Obtener el índice de la última columna
    ultima_columna = ws.max_column

    # Iterar sobre las columnas guardadas y ajustar el tamaño de las columnas a la derecha
    for columna_idx in columnas_dif_vs_py:
        if columna_idx < ultima_columna:
            ws.column_dimensions[get_column_letter(columna_idx + 1)].width = 3  # Ajustar el tamaño de la columna siguiente


    ##COMIENZO a pintar los especiales
    for filas_azul in lista_azul_summary: 
        for fila in ws.iter_rows(min_row=filas_azul, max_row=filas_azul,max_col=num_columnas):
        #if numero_fila not in numeros_filas_activas:
            for celda in fila:
                if celda.value is not None and isinstance(celda.value, (str, int, float)):
                    celda.font = fuente_negrita
                    celda.fill = relleno_azul

    # Aplicar bordes y formatos numéricos
    for numero_fila, fila in enumerate(ws.iter_rows(min_row=1, max_row=num_filas+10, min_col=1, max_col=num_columnas), start=1):
        #if numero_fila not in numeros_filas_activas:
            for celda in fila:
                valor = celda.value
                if  isinstance(valor, str):
                    if valor is not None and valor.strip() == "":
                        celda.border = borde_blanco
                    else:
                        celda.border = borde_negro
                elif isinstance(valor, (int, float)):
                    celda.border = borde_negro


    fuente_roja =  Font(color='FF0000', name='Arial', size=8)
    celda_inferior2 = ws.cell(row=celda_inicio+1, column=1, value=f"Mes: {nombre_del_mes}")
    celda_inferior3 = ws.cell(row=celda_inicio+2, column=1, value="T. Canales - T. Argentina")
    celda_inferior4 = ws.cell(row=celda_inicio+3, column=1, value="* Baja muestra")
    celda_inferior2.font = fuente_negrita
    celda_inferior3.font = fuente_negrita
    celda_inferior4.font = fuente_roja 

    celda_inferior6 = ws.cell(row=celda_inicio+4, column=1, value="Penetración (%)")
    celda_inferior7 = ws.cell(row=celda_inicio+73, column=1, value="Frecuencia (veces)")
    celda_inferior8 = ws.cell(row=celda_inicio+50, column=1, value="Compra por acto (kg)")
    celda_inferior9 = ws.cell(row=celda_inicio+27, column=1, value="Compra media (kg)")
    ##guardar
    wb.save(direccion_base2)
